from sklearn.linear_model import LogisticRegression
from sklearn.metrics import classification_report, roc_curve, confusion_matrix, roc_auc_score
from sklearn.model_selection import GridSearchCV, RandomizedSearchCV
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from keras.preprocessing.image import ImageDataGenerator
from keras.models import Model, load_model, Sequential
from keras.layers.core import Flatten, Dense, Dropout
from keras.layers import Input
from keras.optimizers import SGD, RMSprop, Adam
from keras.callbacks import TensorBoard, EarlyStopping
from keras import backend as K
import keras
import tensorflow as tf
import json
import argparse
import sys
import matplotlib.pyplot as plt
import numpy as np
import os
import pickle
import glob
from models import VGG16
import random
import pandas as pd
import cv2
import gc
from time import time
from openpyxl import load_workbook, Workbook
from utils import *
from sklearn import svm

class NeuralNetwork:
    def __init__(self, model, config, **kwargs):
        self.model = model
        self.config = config
        self.trainingpath = config['trainingpath']
        self.validationpath = config['validationpath']
        self.testpath = config['testpath']
        self.batchsize = kwargs.get('batchsize', 32)
        for attribute, value in kwargs.items():
            setattr(self, attribute, value)

        # initialize image data generator objects
        self.gen_obj_training = ImageDataGenerator(rescale=1./255, featurewise_center=True)
        self.gen_obj_test = ImageDataGenerator(rescale=1./255, featurewise_center=True)

        # fit generators to training data
        x_train = load_training_data(self.trainingpath)
        self.gen_obj_training.fit(x_train, seed=self.seed)
        self.gen_obj_test.fit(x_train, seed=self.seed)

        self.num_training = len(glob.glob(os.path.join(self.trainingpath, '**/*.jpg')))
        self.num_validation = len(glob.glob(os.path.join(self.validationpath, '**/*.jpg')))
        self.num_test= len(glob.glob(os.path.join(self.testpath, '**/*.jpg')))

        # create a savepath for results and create a sheet to avoid errors
        self.savefile = os.path.join(self.config['output_path'], 'results_g.xlsx')

        # also create excel file already to avoid errors, if it doesn't exist yet
        if not os.path.exists(self.savefile):
            Workbook().save(self.savefile)

    def init_generators(self, shuffle_training, shuffle_validation, shuffle_test, **kwargs):
        for attribute, value in kwargs.items():
            setattr(self, attribute, value)

        # initialize image generators that load batches of images
        self.gen_training = self.gen_obj_training.flow_from_directory(
            self.trainingpath,
            class_mode="binary",
            target_size=(224,224),
            color_mode="rgb",
            shuffle=shuffle_training,
            batch_size=self.batchsize)

        self.gen_validation = self.gen_obj_test.flow_from_directory(
            self.validationpath,
            class_mode="binary",
            target_size=(224,224),
            color_mode="rgb",
            shuffle=shuffle_validation,
            batch_size=self.batchsize)

        self.gen_test = self.gen_obj_test.flow_from_directory(
            self.testpath,
            class_mode="binary",
            target_size=(224,224),
            color_mode="rgb",
            shuffle=shuffle_test,
            batch_size=self.batchsize)

    def compile_network(self, learning_rate, **kwargs):
        # compile model
        self.optimizer = kwargs.get('optimizer', 'adam')
        self.loss = kwargs.get('loss', "binary_crossentropy")
        self.metrics = kwargs.get('metrics', ["accuracy"])
        self.learning_rate = learning_rate

        # check which optimizer should be used
        if self.optimizer == 'adam':
            self.optimizer = Adam(lr=self.learning_rate, beta_1=0.9, beta_2=0.999, epsilon=None, decay=0.0, amsgrad=False)
        elif self.optimizer == 'rmsprop':
            self.optimizer = RMSprop(lr=self.learning_rate)
        elif self.optimizer == 'sgd':
            self.optimizer = SGD(lr=self.learning_rate, momentum=0.9, nesterov=True)
        else:
            print("Unsupported optimizer, please choose one of (adam, rmsprop, sgd)")

        self.model.compile(loss=self.loss, optimizer=self.optimizer, metrics=self.metrics)

    def train(self, epochs):
        # initialize image generators
        self.init_generators(shuffle_training=True, shuffle_validation=False, shuffle_test=False, batchsize=32)

        # determine how much time passes
        start = time()

        # early_stopping = EarlyStopping(monitor='val_acc', patience=12, verbose=0, mode='auto', baseline=None, restore_best_weights=True)

        # train the model
        hist = self.model.fit_generator(self.gen_training,
            steps_per_epoch = self.num_training // self.batchsize,
            validation_data = self.gen_validation,
            validation_steps = self.num_validation // self.batchsize,
            epochs=epochs,
            verbose=1)
            # callbacks=[early_stopping])

        # determine how much time passed
        end = time()
        self.training_time = end-start

        return hist.history

    def save_history(self, history):
        # save history
        with pd.ExcelWriter(self.savefile, engine='openpyxl') as writer:
            writer.book = load_workbook(self.savefile)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            pd.DataFrame(history).to_excel(writer, sheet_name='history')

        histsavepath = os.path.join(self.config['model_savepath'], "{}_history_fromscratch".format(args['dataset']))
        with open(histsavepath, 'wb') as file:
            pickle.dump(history, file)

    def save_model(self):
        # save trained model
        savepath = os.path.join(self.config['model_savepath'], "{}_model.h5".format(args['dataset']))
        self.model.save(savepath)

    def plot_training(self, history):
        # get the number of epochs the model was trained for
        epochs = len(history['loss'])

        # plot and save training history
        plt.style.use("ggplot")

        fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=1, sharex=True)

        ax1.set_ylabel('Loss')
        ax1.set_xlim([1,epochs])
        ax1.plot(np.arange(0, epochs), history["loss"], label="train")
        ax1.plot(np.arange(0, epochs), history["val_loss"], label="validation")
        ax1.legend(loc="upper right")

        ax2.set_ylabel('Accuracy')
        ax2.set_ylim([0, 1])
        ax2.plot(np.arange(0, epochs), history["acc"], label="train")
        ax2.plot(np.arange(0, epochs), history["val_acc"], label="validation")
        ax2.legend(loc="lower right")

        # create plot path
        plotpath = os.path.join(self.config['plot_path'], "{}_training.png".format(args['dataset']))

        # save plot
        plt.savefig(plotpath)

    def evaluate(self, **kwargs):
        mode = kwargs.get('mode', 'from_scratch')
        source_dataset = kwargs.get('source_dataset', None)
        savepath = kwargs.get('savepath', os.path.join(self.config['plot_path'], "ROC.png"))
        sksavepath = kwargs.get('sksavepath', os.path.join(self.config['plot_path'], "skROC.png"))

        print("mode: {}".format(mode))

        # initialize image generators
        self.gen_validation.reset()
        self.init_generators(shuffle_training=False, shuffle_validation=False, shuffle_test=False, batchsize=1)

        # make predictions
        preds = self.model.predict_generator(self.gen_validation, verbose=1)

        # preds is an array like [[x] [x] [x]], make it into array like [x x x]
        preds = np.asarray([label for sublist in preds for label in sublist])

        # get true labels
        true_labels = self.gen_validation.classes

        # calculate AUC and sklearn AUC
        fpr, tpr, thresholds, AUC = AUC_score(preds, true_labels)
        skfpr, sktpr, skthresholds, skAUC = skAUC_score(preds, true_labels)

        # calculate accuracy score
        acc = accuracy(preds, true_labels)

        if mode == 'from_scratch':
            # create dataframe out of results
            test_results = pd.DataFrame([{'AUC': AUC, 'skAUC': skAUC, 'acc': acc, 'training_time': self.training_time}])

            # read existing rows and add test results
            try:
                df = pd.read_excel(self.savefile, sheet_name=mode)
            except:
                df = pd.DataFrame(columns=['AUC', 'skAUC', 'acc', 'training_time'])

            df = df.append(test_results, ignore_index=True)
            # df.set_index('skAUC', inplace=True)

            # save results in excel file
            with pd.ExcelWriter(self.savefile, engine='openpyxl') as writer:
                writer.book = load_workbook(self.savefile)
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                # df.index.name = 'skAUC'
                test_results.to_excel(writer, sheet_name='from_scratch')

            # plot AUC plots
            plot_AUC(fpr, tpr, AUC, savepath)
            plot_skAUC(skfpr, sktpr, skAUC, sksavepath)

        if mode in ['fc', 'fine_tuning']:
            # save in path for target dataset
            test_results = pd.Series({'source_dataset': source_dataset, 'AUC': AUC, 'skAUC': skAUC, 'acc': acc, 'training_time': self.training_time})

            # read existing rows and add test results
            try:
                df = pd.read_excel(self.savefile, sheet_name=mode)
            except:
                df = pd.DataFrame(columns=['source_dataset', 'AUC', 'skAUC', 'acc', 'training_time'])

            df = df.append(test_results, ignore_index=True)
            df.set_index('source_dataset', inplace=True)

            # save results in excel file
            with pd.ExcelWriter(self.savefile, engine='openpyxl') as writer:
                writer.book = load_workbook(self.savefile)
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                df.index.name = 'source_dataset'
                df.to_excel(writer, sheet_name=mode)

        return AUC, skAUC, acc

    def set_bottleneck_model(self, outputlayer='flatten_1'):
        # create a bottleneck model until given output layer
        self.model = Model(inputs=self.model.input, outputs=self.model.get_layer(outputlayer).output)

    def extract_bottleneck_features(self):
        # don't shuffle when extracting features
        self.init_generators(shuffle_training=False, shuffle_validation=False, shuffle_test=False, batchsize=1)

        # extract features from bottleneck model
        bn_features_train = self.model.predict_generator(self.gen_training, steps=self.num_training//self.batchsize, verbose=1)
        bn_features_test = self.model.predict_generator(self.gen_test, steps=self.num_test//self.batchsize, verbose=1)

        # get true labels
        true_labels_train = self.gen_training.classes
        true_labels_test = self.gen_test.classes

        return bn_features_train, bn_features_test, true_labels_train, true_labels_test

    def set_ft_model(self):
        # build classification model
        top_model = Dense(128, activation="relu")(self.model.output)
        # top_model = Dropout(0.3)(top_model)
        top_model = Dense(8, activation="relu")(top_model)
        # top_model = Dropout(0.3)(top_model)
        top_model = Dense(1, activation="sigmoid")(top_model)

        # add model on top of base model
        self.model =  Model(inputs=self.model.input, outputs=top_model)
        return self.model

def main():
    # read parameters for wanted dataset from config file
    with open('config.json') as json_file:
        config_json = json.load(json_file)
        config = config_json[args['dataset']]
        if args['mode'] in ['SVM', 'fine_tuning']:
            config_source = config_json[args['source_dataset']]

    # create directories to save results if they don't exist yet
    resultspath = os.path.join(os.path.dirname(os.getcwd()), 'results')
    if not os.path.exists(resultspath):
        os.makedirs(resultspath)

    if not os.path.exists(config['model_savepath']):
        os.makedirs(config['model_savepath'])

    if not os.path.exists(config['plot_path']):
        os.makedirs(config['plot_path'])

    # assign batch size
    batchsize = int(args['batchsize'])

    # set a random seed
    seed=28

    if args['mode'] == 'from_scratch':
        # set random seed for result reproducability
        os.environ['PYTHONHASHSEED'] = str(seed)
        np.random.seed(seed)
        random.seed(seed)
        tf.set_random_seed(seed)
        # tf.random.set_seed(seed)

        # set parameters for training
        learning_rate = 5e-6
        epochs = 100

        # load VGG16 model architecture
        model = VGG16(dropout_rate=0.3, l2_rate=0.0, batchnorm=True, activation='relu', input_shape=(224,224,3)).get_model()
        model.summary()

        # create network instance
        network = NeuralNetwork(model, config, batchsize=batchsize, seed=seed)

        # compile network
        print("compiling network...")
        network.compile_network(learning_rate, optimizer='adam', loss="binary_crossentropy", metrics=["accuracy"])

        # train network
        print("training network...")
        history = network.train(epochs)

        # save history
        print("saving training history...")
        network.save_history(history)

        # save network
        print("saving network...")
        network.save_model()

        # plot training
        print("plotting training progress...")
        network.plot_training(history)

        # evaluate network on test data
        print("evaluating network on test data...")
        network.evaluate(mode=args['mode'])

    if args['mode'] == 'SVM':
        # load the pre-trained source network
        print("loading source network...")
        source_dataset = args['source_dataset']
        modelpath = os.path.join(config_source['model_savepath'], '{}_model.h5'.format(source_dataset))
        source_model = load_model(modelpath)
        source_model.summary()

        # create network instance
        network = NeuralNetwork(source_model, config, batchsize=batchsize, seed=seed)

        # set create a bottleneck model at specified layer
        network.set_bottleneck_model(outputlayer='flatten_1')
        network.model.summary()

        # extract features using bottleneck model
        bn_features_train, bn_features_test, true_labels_train, true_labels_test = network.extract_bottleneck_features()

        # scale the data to zero mean, unit variance for PCA
        scaler = StandardScaler()
        train_features = scaler.fit_transform(bn_features_train)
        test_features = scaler.transform(bn_features_test)

        # determine time taken to perform PCA and train SVM
        start = time()

        # fit PCA
        print("performing PCA...")
        pca = PCA(.95)
        pca.fit(train_features)

        # apply PCA to features and test data
        reduced_train_features = pca.transform(train_features)
        reduced_test_features = pca.transform(test_features)

        # fit SVM classifier
        print("fitting SVM classifier...")
        clf = svm.LinearSVC(penalty='l2', loss='squared_hinge', C=1.0).fit(reduced_train_features, true_labels_train)

        # determine time taken to perform PCA and train SVM
        end = time()
        training_time = end-start

        # make predictions using the trained SVM
        preds = clf.decision_function(reduced_test_features)

        # calculate AUC and sklearn AUC
        print("evaluating results...")
        fpr, tpr, thresholds, AUC = AUC_score(preds, true_labels_test)
        skfpr, sktpr, skthresholds, skAUC = skAUC_score(preds, true_labels_test)

        # calculate accuracy score
        acc = accuracy(preds, true_labels_test)

        # create a savepath for results and create a sheet to avoid errors
        savefile = os.path.join(config['output_path'], 'results_g.xlsx')

        # also create excel file if it doesn't exist yet to avoid errors
        if not os.path.exists(savefile):
            Workbook().save(savefile)

        # save in path for target dataset
        test_results = pd.Series({'source_dataset': source_dataset, 'AUC': AUC, 'skAUC': skAUC, 'acc': acc, 'training_time': training_time})

        # read existing results if they exist, else create empty dataframe
        try:
            df = pd.read_excel(savefile, sheet_name='SVM')
        except:
            df = pd.DataFrame(columns=['source_dataset', 'AUC', 'skAUC', 'acc', 'training_time'])

        df = df.append(test_results, ignore_index=True)
        df.set_index('source_dataset', inplace=True)

        # save results in excel file
        with pd.ExcelWriter(savefile, engine='openpyxl') as writer:
            writer.book = load_workbook(savefile)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            df.index.name = 'source_dataset'
            df.to_excel(writer, sheet_name='SVM')

    if args['mode'] == 'fine_tuning':
        # load the pre-trained source network
        print("loading source network...")
        source_dataset = args['source_dataset']
        modelpath = os.path.join(config_source['model_savepath'], '{}_model.h5'.format(source_dataset))
        source_model = load_model(modelpath)
        source_model.summary()

        # create network instance
        network = NeuralNetwork(source_model, config, batchsize=batchsize, seed=seed)

        # create a bottleneck model at specified layer
        network.set_bottleneck_model(outputlayer='flatten_1')
        network.model.summary()

        # freeze all layers in the convolutional base to exclude them from training
        for layer in network.model.layers:
            layer.trainable = False

        # set learning phase to avoid problems with BN layers freezing/training
        K.set_learning_phase(1)

        # add classification model on top of the bottleneck model
        network.set_ft_model()
        network.model.summary()

        # compile network
        print("compiling network...")
        network.compile_network(learning_rate=1e-3, optimizer='sgd', loss="binary_crossentropy", metrics=["accuracy"])

        # train the model for some epochs
        print("training top model...")
        epochs = 25
        history = network.train(epochs)

        histsavepath = os.path.join(config['model_savepath'], "{}_history_fc".format(args['dataset']))
        with open(histsavepath, 'wb') as file:
            pickle.dump(history, file)

        # evaluate on test data
        print("evaluating top model...")
        network.evaluate(mode='fc', source_dataset=source_dataset)

        # find correct layer index for last conv block and unfreeze last convolutional block
        for idx, layer in enumerate(network.model.layers):
            if layer.name == 'conv2d_11':
                index = idx

        for layer in network.model.layers[index:]:
            layer.trainable = True

        # reset training and validation generators
        network.gen_training.reset()
        network.gen_validation.reset()

        # recompile model
        print("compiling network...")
        network.compile_network(learning_rate=1e-4, optimizer='sgd', loss="binary_crossentropy", metrics=["accuracy"])

        # train model some more
        print("training top model and unfrozen layers...")
        epochs = 25
        history = network.train(epochs)

        histsavepath = os.path.join(config['model_savepath'], "{}_history_finetuning".format(args['dataset']))
        with open(histsavepath, 'wb') as file:
            pickle.dump(history, file)

        # evaluate results on test data
        print("evaluating after fine-tuning top model...")
        network.evaluate(mode='fine_tuning', source_dataset=source_dataset)

    if args['mode'] == 'evaluate':
        # load the source network
        print("loading source network...")
        modelpath = os.path.join(config['model_savepath'], '{}_model.h5'.format(args['dataset']))
        model = load_model(modelpath)
        model.summary()

        # create network instance
        network = NeuralNetwork(model, config, batchsize=batchsize, seed=seed)

        # fit generators on training data
        x_train = load_training_data(network.trainingpath)
        network.gen_obj_training.fit(x_train, seed=seed)
        network.gen_obj_test.fit(x_train, seed=seed)

        # initialize generators
        network.init_generators(shuffle_training=False, shuffle_validation=False, shuffle_test=False, batchsize=1)

        # evaluate network on test data
        print("evaluating network on test data...")
        network.evaluate(mode=args['mode'])


if __name__ == "__main__":
    # choose GPU for training
    # os.environ["CUDA_DEVICE_ORDER"]="PCI_BUS_ID"
    os.environ["CUDA_VISIBLE_DEVICES"] = "2"

    # construct argument parser and parse the arguments
    parser = argparse.ArgumentParser()
    parser.add_argument('-m',
        '--mode',
        choices=['from_scratch', 'SVM', 'fine_tuning', 'evaluate'],
        required=True,
        help='training mode')
    parser.add_argument('-d',
        '--dataset',
        choices=['ISIC_2', 'ISIC_2_image_rot_f=0.1', 'ISIC_2_image_rot_f=0.2',
                'ISIC_2_image_rot_f=0.3', 'ISIC_2_image_rot_f=0.4', 'ISIC_2_image_rot_f=0.5',
                'ISIC_2_image_rot_f=0.6', 'ISIC_2_image_rot_f=0.7', 'ISIC_2_image_rot_f=0.8',
                'ISIC_2_image_rot_f=0.9', 'ISIC_2_image_rot_f=1.0', 'ISIC_2_image_translation_f=0.1',
                'ISIC_2_image_translation_f=0.2', 'ISIC_2_image_translation_f=0.3', 'ISIC_2_image_translation_f=0.4',
                'ISIC_2_image_translation_f=0.5', 'ISIC_2_image_translation_f=0.6', 'ISIC_2_image_translation_f=0.7',
                'ISIC_2_image_translation_f=0.8', 'ISIC_2_image_translation_f=0.9', 'ISIC_2_image_translation_f=1.0',
                'ISIC_2_image_zoom_f=0.1', 'ISIC_2_image_zoom_f=0.2', 'ISIC_2_image_zoom_f=0.3',
                'ISIC_2_image_zoom_f=0.4', 'ISIC_2_image_zoom_f=0.5', 'ISIC_2_image_zoom_f=0.6',
                'ISIC_2_image_zoom_f=0.7', 'ISIC_2_image_zoom_f=0.8', 'ISIC_2_image_zoom_f=0.9',
                'ISIC_2_image_zoom_f=1.0', 'ISIC_2_add_noise_gaussian_f=0.1', 'ISIC_2_add_noise_gaussian_f=0.2',
                'ISIC_2_add_noise_gaussian_f=0.3', 'ISIC_2_add_noise_gaussian_f=0.4', 'ISIC_2_add_noise_gaussian_f=0.5',
                'ISIC_2_add_noise_gaussian_f=0.6', 'ISIC_2_add_noise_gaussian_f=0.7', 'ISIC_2_add_noise_gaussian_f=0.8',
                'ISIC_2_add_noise_gaussian_f=0.9', 'ISIC_2_add_noise_gaussian_f=1.0', 'ISIC_2_add_noise_poisson_f=0.1',
                'ISIC_2_add_noise_poisson_f=0.2', 'ISIC_2_add_noise_poisson_f=0.3', 'ISIC_2_add_noise_poisson_f=0.4',
                'ISIC_2_add_noise_poisson_f=0.5', 'ISIC_2_add_noise_poisson_f=0.6', 'ISIC_2_add_noise_poisson_f=0.7',
                'ISIC_2_add_noise_poisson_f=0.8', 'ISIC_2_add_noise_poisson_f=0.9', 'ISIC_2_add_noise_poisson_f=1.0',
                'ISIC_2_add_noise_salt_and_pepper_f=0.1', 'ISIC_2_add_noise_salt_and_pepper_f=0.2',
                'ISIC_2_add_noise_salt_and_pepper_f=0.3', 'ISIC_2_add_noise_salt_and_pepper_f=0.4',
                'ISIC_2_add_noise_salt_and_pepper_f=0.5', 'ISIC_2_add_noise_salt_and_pepper_f=0.6',
                'ISIC_2_add_noise_salt_and_pepper_f=0.7', 'ISIC_2_add_noise_salt_and_pepper_f=0.8',
                'ISIC_2_add_noise_salt_and_pepper_f=0.9', 'ISIC_2_add_noise_salt_and_pepper_f=1.0',
                'ISIC_2_add_noise_speckle_f=0.1', 'ISIC_2_add_noise_speckle_f=0.2', 'ISIC_2_add_noise_speckle_f=0.3',
                'ISIC_2_add_noise_speckle_f=0.4', 'ISIC_2_add_noise_speckle_f=0.5', 'ISIC_2_add_noise_speckle_f=0.6',
                'ISIC_2_add_noise_speckle_f=0.7', 'ISIC_2_add_noise_speckle_f=0.8', 'ISIC_2_add_noise_speckle_f=0.9',
                'ISIC_2_add_noise_speckle_f=1.0', 'ISIC_2_imbalance_classes_f=0.1', 'ISIC_2_imbalance_classes_f=0.2',
                'ISIC_2_imbalance_classes_f=0.3', 'ISIC_2_imbalance_classes_f=0.4', 'ISIC_2_imbalance_classes_f=0.5',
                'ISIC_2_imbalance_classes_f=0.6', 'ISIC_2_imbalance_classes_f=0.7', 'ISIC_2_imbalance_classes_f=0.8',
                'ISIC_2_imbalance_classes_f=0.9', 'ISIC_2_imbalance_classes_f=1.0', 'ISIC_2_grayscale_f=0.1',
                'ISIC_2_grayscale_f=0.2', 'ISIC_2_grayscale_f=0.3', 'ISIC_2_grayscale_f=0.4',
                'ISIC_2_grayscale_f=0.5', 'ISIC_2_grayscale_f=0.6', 'ISIC_2_grayscale_f=0.7',
                'ISIC_2_grayscale_f=0.8', 'ISIC_2_grayscale_f=0.9', 'ISIC_2_grayscale_f=1.0',
                'ISIC_2_hsv_f=0.1', 'ISIC_2_hsv_f=0.2', 'ISIC_2_hsv_f=0.3', 'ISIC_2_hsv_f=0.4',
                'ISIC_2_hsv_f=0.5', 'ISIC_2_hsv_f=0.6', 'ISIC_2_hsv_f=0.7',
                'ISIC_2_hsv_f=0.8', 'ISIC_2_hsv_f=0.9', 'ISIC_2_hsv_f=1.0',
                'ISIC_2_blur_f=0.1', 'ISIC_2_blur_f=0.2', 'ISIC_2_blur_f=0.3', 'ISIC_2_blur_f=0.4',
                'ISIC_2_blur_f=0.5', 'ISIC_2_blur_f=0.6', 'ISIC_2_blur_f=0.7',
                'ISIC_2_blur_f=0.8', 'ISIC_2_blur_f=0.9', 'ISIC_2_blur_f=1.0',
                'ISIC_2_small_random_f=0.1', 'ISIC_2_small_random_f=0.2', 'ISIC_2_small_random_f=0.3', 'ISIC_2_small_random_f=0.4',
                'ISIC_2_small_random_f=0.5', 'ISIC_2_small_random_f=0.6', 'ISIC_2_small_random_f=0.7',
                'ISIC_2_small_random_f=0.8', 'ISIC_2_small_random_f=0.9', 'ISIC_2_small_random_f=1.0',
                'ISIC_2_small_easy_f=0.1', 'ISIC_2_small_easy_f=0.2', 'ISIC_2_small_easy_f=0.3', 'ISIC_2_small_easy_f=0.4',
                'ISIC_2_small_easy_f=0.5', 'ISIC_2_small_easy_f=0.6', 'ISIC_2_small_easy_f=0.7',
                'ISIC_2_small_easy_f=0.8', 'ISIC_2_small_easy_f=0.9', 'ISIC_2_small_easy_f=1.0',
                'ISIC_2_small_hard_f=0.1', 'ISIC_2_small_hard_f=0.2', 'ISIC_2_small_hard_f=0.3', 'ISIC_2_small_hard_f=0.4',
                'ISIC_2_small_hard_f=0.5', 'ISIC_2_small_hard_f=0.6', 'ISIC_2_small_hard_f=0.7',
                'ISIC_2_small_hard_f=0.8', 'ISIC_2_small_hard_f=0.9', 'ISIC_2_small_hard_f=1.0',
                'ISIC_2_small_clusters_f=0.1', 'ISIC_2_small_clusters_f=0.2', 'ISIC_2_small_clusters_f=0.3', 'ISIC_2_small_clusters_f=0.4',
                'ISIC_2_small_clusters_f=0.5', 'ISIC_2_small_clusters_f=0.6', 'ISIC_2_small_clusters_f=0.7',
                'ISIC_2_small_clusters_f=0.8', 'ISIC_2_small_clusters_f=0.9', 'ISIC_2_small_clusters_f=1.0',
                'ISIC_3', 'ISIC_3_image_rot_f=0.1', 'ISIC_3_image_rot_f=0.2',
                'ISIC_3_image_rot_f=0.3', 'ISIC_3_image_rot_f=0.4', 'ISIC_3_image_rot_f=0.5',
                'ISIC_3_image_rot_f=0.6', 'ISIC_3_image_rot_f=0.7', 'ISIC_3_image_rot_f=0.8',
                'ISIC_3_image_rot_f=0.9', 'ISIC_3_image_rot_f=1.0', 'ISIC_3_image_translation_f=0.1',
                'ISIC_3_image_translation_f=0.2', 'ISIC_3_image_translation_f=0.3', 'ISIC_3_image_translation_f=0.4',
                'ISIC_3_image_translation_f=0.5', 'ISIC_3_image_translation_f=0.6', 'ISIC_3_image_translation_f=0.7',
                'ISIC_3_image_translation_f=0.8', 'ISIC_3_image_translation_f=0.9', 'ISIC_3_image_translation_f=1.0',
                'ISIC_3_image_zoom_f=0.1', 'ISIC_3_image_zoom_f=0.2', 'ISIC_3_image_zoom_f=0.3',
                'ISIC_3_image_zoom_f=0.4', 'ISIC_3_image_zoom_f=0.5', 'ISIC_3_image_zoom_f=0.6',
                'ISIC_3_image_zoom_f=0.7', 'ISIC_3_image_zoom_f=0.8', 'ISIC_3_image_zoom_f=0.9',
                'ISIC_3_image_zoom_f=1.0', 'ISIC_3_add_noise_gaussian_f=0.1', 'ISIC_3_add_noise_gaussian_f=0.2',
                'ISIC_3_add_noise_gaussian_f=0.3', 'ISIC_3_add_noise_gaussian_f=0.4', 'ISIC_3_add_noise_gaussian_f=0.5',
                'ISIC_3_add_noise_gaussian_f=0.6', 'ISIC_3_add_noise_gaussian_f=0.7', 'ISIC_3_add_noise_gaussian_f=0.8',
                'ISIC_3_add_noise_gaussian_f=0.9', 'ISIC_3_add_noise_gaussian_f=1.0', 'ISIC_3_add_noise_poisson_f=0.1',
                'ISIC_3_add_noise_poisson_f=0.2', 'ISIC_3_add_noise_poisson_f=0.3', 'ISIC_3_add_noise_poisson_f=0.4',
                'ISIC_3_add_noise_poisson_f=0.5', 'ISIC_3_add_noise_poisson_f=0.6', 'ISIC_3_add_noise_poisson_f=0.7',
                'ISIC_3_add_noise_poisson_f=0.8', 'ISIC_3_add_noise_poisson_f=0.9', 'ISIC_3_add_noise_poisson_f=1.0',
                'ISIC_3_add_noise_salt_and_pepper_f=0.1', 'ISIC_3_add_noise_salt_and_pepper_f=0.2',
                'ISIC_3_add_noise_salt_and_pepper_f=0.3', 'ISIC_3_add_noise_salt_and_pepper_f=0.4',
                'ISIC_3_add_noise_salt_and_pepper_f=0.5', 'ISIC_3_add_noise_salt_and_pepper_f=0.6',
                'ISIC_3_add_noise_salt_and_pepper_f=0.7', 'ISIC_3_add_noise_salt_and_pepper_f=0.8',
                'ISIC_3_add_noise_salt_and_pepper_f=0.9', 'ISIC_3_add_noise_salt_and_pepper_f=1.0',
                'ISIC_3_add_noise_speckle_f=0.1', 'ISIC_3_add_noise_speckle_f=0.2', 'ISIC_3_add_noise_speckle_f=0.3',
                'ISIC_3_add_noise_speckle_f=0.4', 'ISIC_3_add_noise_speckle_f=0.5', 'ISIC_3_add_noise_speckle_f=0.6',
                'ISIC_3_add_noise_speckle_f=0.7', 'ISIC_3_add_noise_speckle_f=0.8', 'ISIC_3_add_noise_speckle_f=0.9',
                'ISIC_3_add_noise_speckle_f=1.0', 'ISIC_3_imbalance_classes_f=0.1', 'ISIC_3_imbalance_classes_f=0.2',
                'ISIC_3_imbalance_classes_f=0.3', 'ISIC_3_imbalance_classes_f=0.4', 'ISIC_3_imbalance_classes_f=0.5',
                'ISIC_3_imbalance_classes_f=0.6', 'ISIC_3_imbalance_classes_f=0.7', 'ISIC_3_imbalance_classes_f=0.8',
                'ISIC_3_imbalance_classes_f=0.9', 'ISIC_3_imbalance_classes_f=1.0', 'ISIC_3_grayscale_f=0.1',
                'ISIC_3_grayscale_f=0.2', 'ISIC_3_grayscale_f=0.3', 'ISIC_3_grayscale_f=0.4',
                'ISIC_3_grayscale_f=0.5', 'ISIC_3_grayscale_f=0.6', 'ISIC_3_grayscale_f=0.7',
                'ISIC_3_grayscale_f=0.8', 'ISIC_3_grayscale_f=0.9', 'ISIC_3_grayscale_f=1.0',
                'ISIC_3_hsv_f=0.1', 'ISIC_3_hsv_f=0.2', 'ISIC_3_hsv_f=0.3', 'ISIC_3_hsv_f=0.4',
                'ISIC_3_hsv_f=0.5', 'ISIC_3_hsv_f=0.6', 'ISIC_3_hsv_f=0.7',
                'ISIC_3_hsv_f=0.8', 'ISIC_3_hsv_f=0.9', 'ISIC_3_hsv_f=1.0',
                'ISIC_3_blur_f=0.1', 'ISIC_3_blur_f=0.2', 'ISIC_3_blur_f=0.3', 'ISIC_3_blur_f=0.4',
                'ISIC_3_blur_f=0.5', 'ISIC_3_blur_f=0.6', 'ISIC_3_blur_f=0.7',
                'ISIC_3_blur_f=0.8', 'ISIC_3_blur_f=0.9', 'ISIC_3_blur_f=1.0',
                'ISIC_3_small_random_f=0.1', 'ISIC_3_small_random_f=0.2', 'ISIC_3_small_random_f=0.3', 'ISIC_3_small_random_f=0.4',
                'ISIC_3_small_random_f=0.5', 'ISIC_3_small_random_f=0.6', 'ISIC_3_small_random_f=0.7',
                'ISIC_3_small_random_f=0.8', 'ISIC_3_small_random_f=0.9', 'ISIC_3_small_random_f=1.0',
                'ISIC_3_small_easy_f=0.1', 'ISIC_3_small_easy_f=0.2', 'ISIC_3_small_easy_f=0.3', 'ISIC_3_small_easy_f=0.4',
                'ISIC_3_small_easy_f=0.5', 'ISIC_3_small_easy_f=0.6', 'ISIC_3_small_easy_f=0.7',
                'ISIC_3_small_easy_f=0.8', 'ISIC_3_small_easy_f=0.9', 'ISIC_3_small_easy_f=1.0',
                'ISIC_3_small_hard_f=0.1', 'ISIC_3_small_hard_f=0.2', 'ISIC_3_small_hard_f=0.3', 'ISIC_3_small_hard_f=0.4',
                'ISIC_3_small_hard_f=0.5', 'ISIC_3_small_hard_f=0.6', 'ISIC_3_small_hard_f=0.7',
                'ISIC_3_small_hard_f=0.8', 'ISIC_3_small_hard_f=0.9', 'ISIC_3_small_hard_f=1.0',
                'ISIC_3_small_clusters_f=0.1', 'ISIC_3_small_clusters_f=0.2', 'ISIC_3_small_clusters_f=0.3', 'ISIC_3_small_clusters_f=0.4',
                'ISIC_3_small_clusters_f=0.5', 'ISIC_3_small_clusters_f=0.6', 'ISIC_3_small_clusters_f=0.7',
                'ISIC_3_small_clusters_f=0.8', 'ISIC_3_small_clusters_f=0.9', 'ISIC_3_small_clusters_f=1.0',
                'ISIC_4', 'ISIC_4_image_rot_f=0.1', 'ISIC_4_image_rot_f=0.2',
                'ISIC_4_image_rot_f=0.3', 'ISIC_4_image_rot_f=0.4', 'ISIC_4_image_rot_f=0.5',
                'ISIC_4_image_rot_f=0.6', 'ISIC_4_image_rot_f=0.7', 'ISIC_4_image_rot_f=0.8',
                'ISIC_4_image_rot_f=0.9', 'ISIC_4_image_rot_f=1.0', 'ISIC_4_image_translation_f=0.1',
                'ISIC_4_image_translation_f=0.2', 'ISIC_4_image_translation_f=0.3', 'ISIC_4_image_translation_f=0.4',
                'ISIC_4_image_translation_f=0.5', 'ISIC_4_image_translation_f=0.6', 'ISIC_4_image_translation_f=0.7',
                'ISIC_4_image_translation_f=0.8', 'ISIC_4_image_translation_f=0.9', 'ISIC_4_image_translation_f=1.0',
                'ISIC_4_image_zoom_f=0.1', 'ISIC_4_image_zoom_f=0.2', 'ISIC_4_image_zoom_f=0.3',
                'ISIC_4_image_zoom_f=0.4', 'ISIC_4_image_zoom_f=0.5', 'ISIC_4_image_zoom_f=0.6',
                'ISIC_4_image_zoom_f=0.7', 'ISIC_4_image_zoom_f=0.8', 'ISIC_4_image_zoom_f=0.9',
                'ISIC_4_image_zoom_f=1.0', 'ISIC_4_add_noise_gaussian_f=0.1', 'ISIC_4_add_noise_gaussian_f=0.2',
                'ISIC_4_add_noise_gaussian_f=0.3', 'ISIC_4_add_noise_gaussian_f=0.4', 'ISIC_4_add_noise_gaussian_f=0.5',
                'ISIC_4_add_noise_gaussian_f=0.6', 'ISIC_4_add_noise_gaussian_f=0.7', 'ISIC_4_add_noise_gaussian_f=0.8',
                'ISIC_4_add_noise_gaussian_f=0.9', 'ISIC_4_add_noise_gaussian_f=1.0', 'ISIC_4_add_noise_poisson_f=0.1',
                'ISIC_4_add_noise_poisson_f=0.2', 'ISIC_4_add_noise_poisson_f=0.3', 'ISIC_4_add_noise_poisson_f=0.4',
                'ISIC_4_add_noise_poisson_f=0.5', 'ISIC_4_add_noise_poisson_f=0.6', 'ISIC_4_add_noise_poisson_f=0.7',
                'ISIC_4_add_noise_poisson_f=0.8', 'ISIC_4_add_noise_poisson_f=0.9', 'ISIC_4_add_noise_poisson_f=1.0',
                'ISIC_4_add_noise_salt_and_pepper_f=0.1', 'ISIC_4_add_noise_salt_and_pepper_f=0.2',
                'ISIC_4_add_noise_salt_and_pepper_f=0.3', 'ISIC_4_add_noise_salt_and_pepper_f=0.4',
                'ISIC_4_add_noise_salt_and_pepper_f=0.5', 'ISIC_4_add_noise_salt_and_pepper_f=0.6',
                'ISIC_4_add_noise_salt_and_pepper_f=0.7', 'ISIC_4_add_noise_salt_and_pepper_f=0.8',
                'ISIC_4_add_noise_salt_and_pepper_f=0.9', 'ISIC_4_add_noise_salt_and_pepper_f=1.0',
                'ISIC_4_add_noise_speckle_f=0.1', 'ISIC_4_add_noise_speckle_f=0.2', 'ISIC_4_add_noise_speckle_f=0.3',
                'ISIC_4_add_noise_speckle_f=0.4', 'ISIC_4_add_noise_speckle_f=0.5', 'ISIC_4_add_noise_speckle_f=0.6',
                'ISIC_4_add_noise_speckle_f=0.7', 'ISIC_4_add_noise_speckle_f=0.8', 'ISIC_4_add_noise_speckle_f=0.9',
                'ISIC_4_add_noise_speckle_f=1.0', 'ISIC_4_imbalance_classes_f=0.1', 'ISIC_4_imbalance_classes_f=0.2',
                'ISIC_4_imbalance_classes_f=0.3', 'ISIC_4_imbalance_classes_f=0.4', 'ISIC_4_imbalance_classes_f=0.5',
                'ISIC_4_imbalance_classes_f=0.6', 'ISIC_4_imbalance_classes_f=0.7', 'ISIC_4_imbalance_classes_f=0.8',
                'ISIC_4_imbalance_classes_f=0.9', 'ISIC_4_imbalance_classes_f=1.0', 'ISIC_4_grayscale_f=0.1',
                'ISIC_4_grayscale_f=0.2', 'ISIC_4_grayscale_f=0.3', 'ISIC_4_grayscale_f=0.4',
                'ISIC_4_grayscale_f=0.5', 'ISIC_4_grayscale_f=0.6', 'ISIC_4_grayscale_f=0.7',
                'ISIC_4_grayscale_f=0.8', 'ISIC_4_grayscale_f=0.9', 'ISIC_4_grayscale_f=1.0',
                'ISIC_4_hsv_f=0.1', 'ISIC_4_hsv_f=0.2', 'ISIC_4_hsv_f=0.3', 'ISIC_4_hsv_f=0.4',
                'ISIC_4_hsv_f=0.5', 'ISIC_4_hsv_f=0.6', 'ISIC_4_hsv_f=0.7',
                'ISIC_4_hsv_f=0.8', 'ISIC_4_hsv_f=0.9', 'ISIC_4_hsv_f=1.0',
                'ISIC_4_blur_f=0.1', 'ISIC_4_blur_f=0.2', 'ISIC_4_blur_f=0.3', 'ISIC_4_blur_f=0.4',
                'ISIC_4_blur_f=0.5', 'ISIC_4_blur_f=0.6', 'ISIC_4_blur_f=0.7',
                'ISIC_4_blur_f=0.8', 'ISIC_4_blur_f=0.9', 'ISIC_4_blur_f=1.0',
                'ISIC_4_small_random_f=0.1', 'ISIC_4_small_random_f=0.2', 'ISIC_4_small_random_f=0.3', 'ISIC_4_small_random_f=0.4',
                'ISIC_4_small_random_f=0.5', 'ISIC_4_small_random_f=0.6', 'ISIC_4_small_random_f=0.7',
                'ISIC_4_small_random_f=0.8', 'ISIC_4_small_random_f=0.9', 'ISIC_4_small_random_f=1.0',
                'ISIC_4_small_easy_f=0.1', 'ISIC_4_small_easy_f=0.2', 'ISIC_4_small_easy_f=0.3', 'ISIC_4_small_easy_f=0.4',
                'ISIC_4_small_easy_f=0.5', 'ISIC_4_small_easy_f=0.6', 'ISIC_4_small_easy_f=0.7',
                'ISIC_4_small_easy_f=0.8', 'ISIC_4_small_easy_f=0.9', 'ISIC_4_small_easy_f=1.0',
                'ISIC_4_small_hard_f=0.1', 'ISIC_4_small_hard_f=0.2', 'ISIC_4_small_hard_f=0.3', 'ISIC_4_small_hard_f=0.4',
                'ISIC_4_small_hard_f=0.5', 'ISIC_4_small_hard_f=0.6', 'ISIC_4_small_hard_f=0.7',
                'ISIC_4_small_hard_f=0.8', 'ISIC_4_small_hard_f=0.9', 'ISIC_4_small_hard_f=1.0',
                'ISIC_4_small_clusters_f=0.1', 'ISIC_4_small_clusters_f=0.2', 'ISIC_4_small_clusters_f=0.3', 'ISIC_4_small_clusters_f=0.4',
                'ISIC_4_small_clusters_f=0.5', 'ISIC_4_small_clusters_f=0.6', 'ISIC_4_small_clusters_f=0.7',
                'ISIC_4_small_clusters_f=0.8', 'ISIC_4_small_clusters_f=0.9', 'ISIC_4_small_clusters_f=1.0',
                'ISIC_5', 'ISIC_5_image_rot_f=0.1', 'ISIC_5_image_rot_f=0.2',
                'ISIC_5_image_rot_f=0.3', 'ISIC_5_image_rot_f=0.4', 'ISIC_5_image_rot_f=0.5',
                'ISIC_5_image_rot_f=0.6', 'ISIC_5_image_rot_f=0.7', 'ISIC_5_image_rot_f=0.8',
                'ISIC_5_image_rot_f=0.9', 'ISIC_5_image_rot_f=1.0', 'ISIC_5_image_translation_f=0.1',
                'ISIC_5_image_translation_f=0.2', 'ISIC_5_image_translation_f=0.3', 'ISIC_5_image_translation_f=0.4',
                'ISIC_5_image_translation_f=0.5', 'ISIC_5_image_translation_f=0.6', 'ISIC_5_image_translation_f=0.7',
                'ISIC_5_image_translation_f=0.8', 'ISIC_5_image_translation_f=0.9', 'ISIC_5_image_translation_f=1.0',
                'ISIC_5_image_zoom_f=0.1', 'ISIC_5_image_zoom_f=0.2', 'ISIC_5_image_zoom_f=0.3',
                'ISIC_5_image_zoom_f=0.4', 'ISIC_5_image_zoom_f=0.5', 'ISIC_5_image_zoom_f=0.6',
                'ISIC_5_image_zoom_f=0.7', 'ISIC_5_image_zoom_f=0.8', 'ISIC_5_image_zoom_f=0.9',
                'ISIC_5_image_zoom_f=1.0', 'ISIC_5_add_noise_gaussian_f=0.1', 'ISIC_5_add_noise_gaussian_f=0.2',
                'ISIC_5_add_noise_gaussian_f=0.3', 'ISIC_5_add_noise_gaussian_f=0.4', 'ISIC_5_add_noise_gaussian_f=0.5',
                'ISIC_5_add_noise_gaussian_f=0.6', 'ISIC_5_add_noise_gaussian_f=0.7', 'ISIC_5_add_noise_gaussian_f=0.8',
                'ISIC_5_add_noise_gaussian_f=0.9', 'ISIC_5_add_noise_gaussian_f=1.0', 'ISIC_5_add_noise_poisson_f=0.1',
                'ISIC_5_add_noise_poisson_f=0.2', 'ISIC_5_add_noise_poisson_f=0.3', 'ISIC_5_add_noise_poisson_f=0.4',
                'ISIC_5_add_noise_poisson_f=0.5', 'ISIC_5_add_noise_poisson_f=0.6', 'ISIC_5_add_noise_poisson_f=0.7',
                'ISIC_5_add_noise_poisson_f=0.8', 'ISIC_5_add_noise_poisson_f=0.9', 'ISIC_5_add_noise_poisson_f=1.0',
                'ISIC_5_add_noise_salt_and_pepper_f=0.1', 'ISIC_5_add_noise_salt_and_pepper_f=0.2',
                'ISIC_5_add_noise_salt_and_pepper_f=0.3', 'ISIC_5_add_noise_salt_and_pepper_f=0.4',
                'ISIC_5_add_noise_salt_and_pepper_f=0.5', 'ISIC_5_add_noise_salt_and_pepper_f=0.6',
                'ISIC_5_add_noise_salt_and_pepper_f=0.7', 'ISIC_5_add_noise_salt_and_pepper_f=0.8',
                'ISIC_5_add_noise_salt_and_pepper_f=0.9', 'ISIC_5_add_noise_salt_and_pepper_f=1.0',
                'ISIC_5_add_noise_speckle_f=0.1', 'ISIC_5_add_noise_speckle_f=0.2', 'ISIC_5_add_noise_speckle_f=0.3',
                'ISIC_5_add_noise_speckle_f=0.4', 'ISIC_5_add_noise_speckle_f=0.5', 'ISIC_5_add_noise_speckle_f=0.6',
                'ISIC_5_add_noise_speckle_f=0.7', 'ISIC_5_add_noise_speckle_f=0.8', 'ISIC_5_add_noise_speckle_f=0.9',
                'ISIC_5_add_noise_speckle_f=1.0', 'ISIC_5_imbalance_classes_f=0.1', 'ISIC_5_imbalance_classes_f=0.2',
                'ISIC_5_imbalance_classes_f=0.3', 'ISIC_5_imbalance_classes_f=0.4', 'ISIC_5_imbalance_classes_f=0.5',
                'ISIC_5_imbalance_classes_f=0.6', 'ISIC_5_imbalance_classes_f=0.7', 'ISIC_5_imbalance_classes_f=0.8',
                'ISIC_5_imbalance_classes_f=0.9', 'ISIC_5_imbalance_classes_f=1.0', 'ISIC_5_grayscale_f=0.1',
                'ISIC_5_grayscale_f=0.2', 'ISIC_5_grayscale_f=0.3', 'ISIC_5_grayscale_f=0.4',
                'ISIC_5_grayscale_f=0.5', 'ISIC_5_grayscale_f=0.6', 'ISIC_5_grayscale_f=0.7',
                'ISIC_5_grayscale_f=0.8', 'ISIC_5_grayscale_f=0.9', 'ISIC_5_grayscale_f=1.0',
                'ISIC_5_hsv_f=0.1', 'ISIC_5_hsv_f=0.2', 'ISIC_5_hsv_f=0.3', 'ISIC_5_hsv_f=0.4',
                'ISIC_5_hsv_f=0.5', 'ISIC_5_hsv_f=0.6', 'ISIC_5_hsv_f=0.7',
                'ISIC_5_hsv_f=0.8', 'ISIC_5_hsv_f=0.9', 'ISIC_5_hsv_f=1.0',
                'ISIC_5_blur_f=0.1', 'ISIC_5_blur_f=0.2', 'ISIC_5_blur_f=0.3', 'ISIC_5_blur_f=0.4',
                'ISIC_5_blur_f=0.5', 'ISIC_5_blur_f=0.6', 'ISIC_5_blur_f=0.7',
                'ISIC_5_blur_f=0.8', 'ISIC_5_blur_f=0.9', 'ISIC_5_blur_f=1.0',
                'ISIC_5_small_random_f=0.1', 'ISIC_5_small_random_f=0.2', 'ISIC_5_small_random_f=0.3', 'ISIC_5_small_random_f=0.4',
                'ISIC_5_small_random_f=0.5', 'ISIC_5_small_random_f=0.6', 'ISIC_5_small_random_f=0.7',
                'ISIC_5_small_random_f=0.8', 'ISIC_5_small_random_f=0.9', 'ISIC_5_small_random_f=1.0',
                'ISIC_5_small_easy_f=0.1', 'ISIC_5_small_easy_f=0.2', 'ISIC_5_small_easy_f=0.3', 'ISIC_5_small_easy_f=0.4',
                'ISIC_5_small_easy_f=0.5', 'ISIC_5_small_easy_f=0.6', 'ISIC_5_small_easy_f=0.7',
                'ISIC_5_small_easy_f=0.8', 'ISIC_5_small_easy_f=0.9', 'ISIC_5_small_easy_f=1.0',
                'ISIC_5_small_hard_f=0.1', 'ISIC_5_small_hard_f=0.2', 'ISIC_5_small_hard_f=0.3', 'ISIC_5_small_hard_f=0.4',
                'ISIC_5_small_hard_f=0.5', 'ISIC_5_small_hard_f=0.6', 'ISIC_5_small_hard_f=0.7',
                'ISIC_5_small_hard_f=0.8', 'ISIC_5_small_hard_f=0.9', 'ISIC_5_small_hard_f=1.0',
                'ISIC_5_small_clusters_f=0.1', 'ISIC_5_small_clusters_f=0.2', 'ISIC_5_small_clusters_f=0.3', 'ISIC_5_small_clusters_f=0.4',
                'ISIC_5_small_clusters_f=0.5', 'ISIC_5_small_clusters_f=0.6', 'ISIC_5_small_clusters_f=0.7',
                'ISIC_5_small_clusters_f=0.8', 'ISIC_5_small_clusters_f=0.9', 'ISIC_5_small_clusters_f=1.0',
                'ISIC_6', 'ISIC_6_image_rot_f=0.1', 'ISIC_6_image_rot_f=0.2',
                'ISIC_6_image_rot_f=0.3', 'ISIC_6_image_rot_f=0.4', 'ISIC_6_image_rot_f=0.5',
                'ISIC_6_image_rot_f=0.6', 'ISIC_6_image_rot_f=0.7', 'ISIC_6_image_rot_f=0.8',
                'ISIC_6_image_rot_f=0.9', 'ISIC_6_image_rot_f=1.0', 'ISIC_6_image_translation_f=0.1',
                'ISIC_6_image_translation_f=0.2', 'ISIC_6_image_translation_f=0.3', 'ISIC_6_image_translation_f=0.4',
                'ISIC_6_image_translation_f=0.5', 'ISIC_6_image_translation_f=0.6', 'ISIC_6_image_translation_f=0.7',
                'ISIC_6_image_translation_f=0.8', 'ISIC_6_image_translation_f=0.9', 'ISIC_6_image_translation_f=1.0',
                'ISIC_6_image_zoom_f=0.1', 'ISIC_6_image_zoom_f=0.2', 'ISIC_6_image_zoom_f=0.3',
                'ISIC_6_image_zoom_f=0.4', 'ISIC_6_image_zoom_f=0.5', 'ISIC_6_image_zoom_f=0.6',
                'ISIC_6_image_zoom_f=0.7', 'ISIC_6_image_zoom_f=0.8', 'ISIC_6_image_zoom_f=0.9',
                'ISIC_6_image_zoom_f=1.0', 'ISIC_6_add_noise_gaussian_f=0.1', 'ISIC_6_add_noise_gaussian_f=0.2',
                'ISIC_6_add_noise_gaussian_f=0.3', 'ISIC_6_add_noise_gaussian_f=0.4', 'ISIC_6_add_noise_gaussian_f=0.5',
                'ISIC_6_add_noise_gaussian_f=0.6', 'ISIC_6_add_noise_gaussian_f=0.7', 'ISIC_6_add_noise_gaussian_f=0.8',
                'ISIC_6_add_noise_gaussian_f=0.9', 'ISIC_6_add_noise_gaussian_f=1.0', 'ISIC_6_add_noise_poisson_f=0.1',
                'ISIC_6_add_noise_poisson_f=0.2', 'ISIC_6_add_noise_poisson_f=0.3', 'ISIC_6_add_noise_poisson_f=0.4',
                'ISIC_6_add_noise_poisson_f=0.5', 'ISIC_6_add_noise_poisson_f=0.6', 'ISIC_6_add_noise_poisson_f=0.7',
                'ISIC_6_add_noise_poisson_f=0.8', 'ISIC_6_add_noise_poisson_f=0.9', 'ISIC_6_add_noise_poisson_f=1.0',
                'ISIC_6_add_noise_salt_and_pepper_f=0.1', 'ISIC_6_add_noise_salt_and_pepper_f=0.2',
                'ISIC_6_add_noise_salt_and_pepper_f=0.3', 'ISIC_6_add_noise_salt_and_pepper_f=0.4',
                'ISIC_6_add_noise_salt_and_pepper_f=0.5', 'ISIC_6_add_noise_salt_and_pepper_f=0.6',
                'ISIC_6_add_noise_salt_and_pepper_f=0.7', 'ISIC_6_add_noise_salt_and_pepper_f=0.8',
                'ISIC_6_add_noise_salt_and_pepper_f=0.9', 'ISIC_6_add_noise_salt_and_pepper_f=1.0',
                'ISIC_6_add_noise_speckle_f=0.1', 'ISIC_6_add_noise_speckle_f=0.2', 'ISIC_6_add_noise_speckle_f=0.3',
                'ISIC_6_add_noise_speckle_f=0.4', 'ISIC_6_add_noise_speckle_f=0.5', 'ISIC_6_add_noise_speckle_f=0.6',
                'ISIC_6_add_noise_speckle_f=0.7', 'ISIC_6_add_noise_speckle_f=0.8', 'ISIC_6_add_noise_speckle_f=0.9',
                'ISIC_6_add_noise_speckle_f=1.0', 'ISIC_6_imbalance_classes_f=0.1', 'ISIC_6_imbalance_classes_f=0.2',
                'ISIC_6_imbalance_classes_f=0.3', 'ISIC_6_imbalance_classes_f=0.4', 'ISIC_6_imbalance_classes_f=0.5',
                'ISIC_6_imbalance_classes_f=0.6', 'ISIC_6_imbalance_classes_f=0.7', 'ISIC_6_imbalance_classes_f=0.8',
                'ISIC_6_imbalance_classes_f=0.9', 'ISIC_6_imbalance_classes_f=1.0', 'ISIC_6_grayscale_f=0.1',
                'ISIC_6_grayscale_f=0.2', 'ISIC_6_grayscale_f=0.3', 'ISIC_6_grayscale_f=0.4',
                'ISIC_6_grayscale_f=0.5', 'ISIC_6_grayscale_f=0.6', 'ISIC_6_grayscale_f=0.7',
                'ISIC_6_grayscale_f=0.8', 'ISIC_6_grayscale_f=0.9', 'ISIC_6_grayscale_f=1.0',
                'ISIC_6_hsv_f=0.1', 'ISIC_6_hsv_f=0.2', 'ISIC_6_hsv_f=0.3', 'ISIC_6_hsv_f=0.4',
                'ISIC_6_hsv_f=0.5', 'ISIC_6_hsv_f=0.6', 'ISIC_6_hsv_f=0.7',
                'ISIC_6_hsv_f=0.8', 'ISIC_6_hsv_f=0.9', 'ISIC_6_hsv_f=1.0',
                'ISIC_6_blur_f=0.1', 'ISIC_6_blur_f=0.2', 'ISIC_6_blur_f=0.3', 'ISIC_6_blur_f=0.4',
                'ISIC_6_blur_f=0.5', 'ISIC_6_blur_f=0.6', 'ISIC_6_blur_f=0.7',
                'ISIC_6_blur_f=0.8', 'ISIC_6_blur_f=0.9', 'ISIC_6_blur_f=1.0',
                'ISIC_6_small_random_f=0.1', 'ISIC_6_small_random_f=0.2', 'ISIC_6_small_random_f=0.3', 'ISIC_6_small_random_f=0.4',
                'ISIC_6_small_random_f=0.5', 'ISIC_6_small_random_f=0.6', 'ISIC_6_small_random_f=0.7',
                'ISIC_6_small_random_f=0.8', 'ISIC_6_small_random_f=0.9', 'ISIC_6_small_random_f=1.0',
                'ISIC_6_small_easy_f=0.1', 'ISIC_6_small_easy_f=0.2', 'ISIC_6_small_easy_f=0.3', 'ISIC_6_small_easy_f=0.4',
                'ISIC_6_small_easy_f=0.5', 'ISIC_6_small_easy_f=0.6', 'ISIC_6_small_easy_f=0.7',
                'ISIC_6_small_easy_f=0.8', 'ISIC_6_small_easy_f=0.9', 'ISIC_6_small_easy_f=1.0',
                'ISIC_6_small_hard_f=0.1', 'ISIC_6_small_hard_f=0.2', 'ISIC_6_small_hard_f=0.3', 'ISIC_6_small_hard_f=0.4',
                'ISIC_6_small_hard_f=0.5', 'ISIC_6_small_hard_f=0.6', 'ISIC_6_small_hard_f=0.7',
                'ISIC_6_small_hard_f=0.8', 'ISIC_6_small_hard_f=0.9', 'ISIC_6_small_hard_f=1.0',
                'ISIC_6_small_clusters_f=0.1', 'ISIC_6_small_clusters_f=0.2', 'ISIC_6_small_clusters_f=0.3', 'ISIC_6_small_clusters_f=0.4',
                'ISIC_6_small_clusters_f=0.5', 'ISIC_6_small_clusters_f=0.6', 'ISIC_6_small_clusters_f=0.7',
                'ISIC_6_small_clusters_f=0.8', 'ISIC_6_small_clusters_f=0.9', 'ISIC_6_small_clusters_f=1.0',
                'CNMC_2', 'CNMC_2_image_rot_f=0.1', 'CNMC_2_image_rot_f=0.2',
                'CNMC_2_image_rot_f=0.3', 'CNMC_2_image_rot_f=0.4', 'CNMC_2_image_rot_f=0.5',
                'CNMC_2_image_rot_f=0.6', 'CNMC_2_image_rot_f=0.7', 'CNMC_2_image_rot_f=0.8',
                'CNMC_2_image_rot_f=0.9', 'CNMC_2_image_rot_f=1.0', 'CNMC_2_image_translation_f=0.1',
                'CNMC_2_image_translation_f=0.2', 'CNMC_2_image_translation_f=0.3', 'CNMC_2_image_translation_f=0.4',
                'CNMC_2_image_translation_f=0.5', 'CNMC_2_image_translation_f=0.6', 'CNMC_2_image_translation_f=0.7',
                'CNMC_2_image_translation_f=0.8', 'CNMC_2_image_translation_f=0.9', 'CNMC_2_image_translation_f=1.0',
                'CNMC_2_image_zoom_f=0.1', 'CNMC_2_image_zoom_f=0.2', 'CNMC_2_image_zoom_f=0.3',
                'CNMC_2_image_zoom_f=0.4', 'CNMC_2_image_zoom_f=0.5', 'CNMC_2_image_zoom_f=0.6',
                'CNMC_2_image_zoom_f=0.7', 'CNMC_2_image_zoom_f=0.8', 'CNMC_2_image_zoom_f=0.9',
                'CNMC_2_image_zoom_f=1.0', 'CNMC_2_add_noise_gaussian_f=0.1', 'CNMC_2_add_noise_gaussian_f=0.2',
                'CNMC_2_add_noise_gaussian_f=0.3', 'CNMC_2_add_noise_gaussian_f=0.4', 'CNMC_2_add_noise_gaussian_f=0.5',
                'CNMC_2_add_noise_gaussian_f=0.6', 'CNMC_2_add_noise_gaussian_f=0.7', 'CNMC_2_add_noise_gaussian_f=0.8',
                'CNMC_2_add_noise_gaussian_f=0.9', 'CNMC_2_add_noise_gaussian_f=1.0', 'CNMC_2_add_noise_poisson_f=0.1',
                'CNMC_2_add_noise_poisson_f=0.2', 'CNMC_2_add_noise_poisson_f=0.3', 'CNMC_2_add_noise_poisson_f=0.4',
                'CNMC_2_add_noise_poisson_f=0.5', 'CNMC_2_add_noise_poisson_f=0.6', 'CNMC_2_add_noise_poisson_f=0.7',
                'CNMC_2_add_noise_poisson_f=0.8', 'CNMC_2_add_noise_poisson_f=0.9', 'CNMC_2_add_noise_poisson_f=1.0',
                'CNMC_2_add_noise_salt_and_pepper_f=0.1', 'CNMC_2_add_noise_salt_and_pepper_f=0.2',
                'CNMC_2_add_noise_salt_and_pepper_f=0.3', 'CNMC_2_add_noise_salt_and_pepper_f=0.4',
                'CNMC_2_add_noise_salt_and_pepper_f=0.5', 'CNMC_2_add_noise_salt_and_pepper_f=0.6',
                'CNMC_2_add_noise_salt_and_pepper_f=0.7', 'CNMC_2_add_noise_salt_and_pepper_f=0.8',
                'CNMC_2_add_noise_salt_and_pepper_f=0.9', 'CNMC_2_add_noise_salt_and_pepper_f=1.0',
                'CNMC_2_add_noise_speckle_f=0.1', 'CNMC_2_add_noise_speckle_f=0.2', 'CNMC_2_add_noise_speckle_f=0.3',
                'CNMC_2_add_noise_speckle_f=0.4', 'CNMC_2_add_noise_speckle_f=0.5', 'CNMC_2_add_noise_speckle_f=0.6',
                'CNMC_2_add_noise_speckle_f=0.7', 'CNMC_2_add_noise_speckle_f=0.8', 'CNMC_2_add_noise_speckle_f=0.9',
                'CNMC_2_add_noise_speckle_f=1.0', 'CNMC_2_imbalance_classes_f=0.1', 'CNMC_2_imbalance_classes_f=0.2',
                'CNMC_2_imbalance_classes_f=0.3', 'CNMC_2_imbalance_classes_f=0.4', 'CNMC_2_imbalance_classes_f=0.5',
                'CNMC_2_imbalance_classes_f=0.6', 'CNMC_2_imbalance_classes_f=0.7', 'CNMC_2_imbalance_classes_f=0.8',
                'CNMC_2_imbalance_classes_f=0.9', 'CNMC_2_imbalance_classes_f=1.0', 'CNMC_2_grayscale_f=0.1',
                'CNMC_2_grayscale_f=0.2', 'CNMC_2_grayscale_f=0.3', 'CNMC_2_grayscale_f=0.4',
                'CNMC_2_grayscale_f=0.5', 'CNMC_2_grayscale_f=0.6', 'CNMC_2_grayscale_f=0.7',
                'CNMC_2_grayscale_f=0.8', 'CNMC_2_grayscale_f=0.9', 'CNMC_2_grayscale_f=1.0',
                'CNMC_2_hsv_f=0.1', 'CNMC_2_hsv_f=0.2', 'CNMC_2_hsv_f=0.3', 'CNMC_2_hsv_f=0.4',
                'CNMC_2_hsv_f=0.5', 'CNMC_2_hsv_f=0.6', 'CNMC_2_hsv_f=0.7',
                'CNMC_2_hsv_f=0.8', 'CNMC_2_hsv_f=0.9', 'CNMC_2_hsv_f=1.0',
                'CNMC_2_blur_f=0.1', 'CNMC_2_blur_f=0.2', 'CNMC_2_blur_f=0.3', 'CNMC_2_blur_f=0.4',
                'CNMC_2_blur_f=0.5', 'CNMC_2_blur_f=0.6', 'CNMC_2_blur_f=0.7',
                'CNMC_2_blur_f=0.8', 'CNMC_2_blur_f=0.9', 'CNMC_2_blur_f=1.0',
                'CNMC_2_small_random_f=0.1', 'CNMC_2_small_random_f=0.2', 'CNMC_2_small_random_f=0.3', 'CNMC_2_small_random_f=0.4',
                'CNMC_2_small_random_f=0.5', 'CNMC_2_small_random_f=0.6', 'CNMC_2_small_random_f=0.7',
                'CNMC_2_small_random_f=0.8', 'CNMC_2_small_random_f=0.9', 'CNMC_2_small_random_f=1.0',
                'CNMC_2_small_easy_f=0.1', 'CNMC_2_small_easy_f=0.2', 'CNMC_2_small_easy_f=0.3', 'CNMC_2_small_easy_f=0.4',
                'CNMC_2_small_easy_f=0.5', 'CNMC_2_small_easy_f=0.6', 'CNMC_2_small_easy_f=0.7',
                'CNMC_2_small_easy_f=0.8', 'CNMC_2_small_easy_f=0.9', 'CNMC_2_small_easy_f=1.0',
                'CNMC_2_small_hard_f=0.1', 'CNMC_2_small_hard_f=0.2', 'CNMC_2_small_hard_f=0.3', 'CNMC_2_small_hard_f=0.4',
                'CNMC_2_small_hard_f=0.5', 'CNMC_2_small_hard_f=0.6', 'CNMC_2_small_hard_f=0.7',
                'CNMC_2_small_hard_f=0.8', 'CNMC_2_small_hard_f=0.9', 'CNMC_2_small_hard_f=1.0',
                'CNMC_2_small_clusters_f=0.1', 'CNMC_2_small_clusters_f=0.2', 'CNMC_2_small_clusters_f=0.3', 'CNMC_2_small_clusters_f=0.4',
                'CNMC_2_small_clusters_f=0.5', 'CNMC_2_small_clusters_f=0.6', 'CNMC_2_small_clusters_f=0.7',
                'CNMC_2_small_clusters_f=0.8', 'CNMC_2_small_clusters_f=0.9', 'CNMC_2_small_clusters_f=1.0',
                'CNMC_3', 'CNMC_3_image_rot_f=0.1', 'CNMC_3_image_rot_f=0.2',
                'CNMC_3_image_rot_f=0.3', 'CNMC_3_image_rot_f=0.4', 'CNMC_3_image_rot_f=0.5',
                'CNMC_3_image_rot_f=0.6', 'CNMC_3_image_rot_f=0.7', 'CNMC_3_image_rot_f=0.8',
                'CNMC_3_image_rot_f=0.9', 'CNMC_3_image_rot_f=1.0', 'CNMC_3_image_translation_f=0.1',
                'CNMC_3_image_translation_f=0.2', 'CNMC_3_image_translation_f=0.3', 'CNMC_3_image_translation_f=0.4',
                'CNMC_3_image_translation_f=0.5', 'CNMC_3_image_translation_f=0.6', 'CNMC_3_image_translation_f=0.7',
                'CNMC_3_image_translation_f=0.8', 'CNMC_3_image_translation_f=0.9', 'CNMC_3_image_translation_f=1.0',
                'CNMC_3_image_zoom_f=0.1', 'CNMC_3_image_zoom_f=0.2', 'CNMC_3_image_zoom_f=0.3',
                'CNMC_3_image_zoom_f=0.4', 'CNMC_3_image_zoom_f=0.5', 'CNMC_3_image_zoom_f=0.6',
                'CNMC_3_image_zoom_f=0.7', 'CNMC_3_image_zoom_f=0.8', 'CNMC_3_image_zoom_f=0.9',
                'CNMC_3_image_zoom_f=1.0', 'CNMC_3_add_noise_gaussian_f=0.1', 'CNMC_3_add_noise_gaussian_f=0.2',
                'CNMC_3_add_noise_gaussian_f=0.3', 'CNMC_3_add_noise_gaussian_f=0.4', 'CNMC_3_add_noise_gaussian_f=0.5',
                'CNMC_3_add_noise_gaussian_f=0.6', 'CNMC_3_add_noise_gaussian_f=0.7', 'CNMC_3_add_noise_gaussian_f=0.8',
                'CNMC_3_add_noise_gaussian_f=0.9', 'CNMC_3_add_noise_gaussian_f=1.0', 'CNMC_3_add_noise_poisson_f=0.1',
                'CNMC_3_add_noise_poisson_f=0.2', 'CNMC_3_add_noise_poisson_f=0.3', 'CNMC_3_add_noise_poisson_f=0.4',
                'CNMC_3_add_noise_poisson_f=0.5', 'CNMC_3_add_noise_poisson_f=0.6', 'CNMC_3_add_noise_poisson_f=0.7',
                'CNMC_3_add_noise_poisson_f=0.8', 'CNMC_3_add_noise_poisson_f=0.9', 'CNMC_3_add_noise_poisson_f=1.0',
                'CNMC_3_add_noise_salt_and_pepper_f=0.1', 'CNMC_3_add_noise_salt_and_pepper_f=0.2',
                'CNMC_3_add_noise_salt_and_pepper_f=0.3', 'CNMC_3_add_noise_salt_and_pepper_f=0.4',
                'CNMC_3_add_noise_salt_and_pepper_f=0.5', 'CNMC_3_add_noise_salt_and_pepper_f=0.6',
                'CNMC_3_add_noise_salt_and_pepper_f=0.7', 'CNMC_3_add_noise_salt_and_pepper_f=0.8',
                'CNMC_3_add_noise_salt_and_pepper_f=0.9', 'CNMC_3_add_noise_salt_and_pepper_f=1.0',
                'CNMC_3_add_noise_speckle_f=0.1', 'CNMC_3_add_noise_speckle_f=0.2', 'CNMC_3_add_noise_speckle_f=0.3',
                'CNMC_3_add_noise_speckle_f=0.4', 'CNMC_3_add_noise_speckle_f=0.5', 'CNMC_3_add_noise_speckle_f=0.6',
                'CNMC_3_add_noise_speckle_f=0.7', 'CNMC_3_add_noise_speckle_f=0.8', 'CNMC_3_add_noise_speckle_f=0.9',
                'CNMC_3_add_noise_speckle_f=1.0', 'CNMC_3_imbalance_classes_f=0.1', 'CNMC_3_imbalance_classes_f=0.2',
                'CNMC_3_imbalance_classes_f=0.3', 'CNMC_3_imbalance_classes_f=0.4', 'CNMC_3_imbalance_classes_f=0.5',
                'CNMC_3_imbalance_classes_f=0.6', 'CNMC_3_imbalance_classes_f=0.7', 'CNMC_3_imbalance_classes_f=0.8',
                'CNMC_3_imbalance_classes_f=0.9', 'CNMC_3_imbalance_classes_f=1.0', 'CNMC_3_grayscale_f=0.1',
                'CNMC_3_grayscale_f=0.2', 'CNMC_3_grayscale_f=0.3', 'CNMC_3_grayscale_f=0.4',
                'CNMC_3_grayscale_f=0.5', 'CNMC_3_grayscale_f=0.6', 'CNMC_3_grayscale_f=0.7',
                'CNMC_3_grayscale_f=0.8', 'CNMC_3_grayscale_f=0.9', 'CNMC_3_grayscale_f=1.0',
                'CNMC_3_hsv_f=0.1', 'CNMC_3_hsv_f=0.2', 'CNMC_3_hsv_f=0.3', 'CNMC_3_hsv_f=0.4',
                'CNMC_3_hsv_f=0.5', 'CNMC_3_hsv_f=0.6', 'CNMC_3_hsv_f=0.7',
                'CNMC_3_hsv_f=0.8', 'CNMC_3_hsv_f=0.9', 'CNMC_3_hsv_f=1.0',
                'CNMC_3_blur_f=0.1', 'CNMC_3_blur_f=0.2', 'CNMC_3_blur_f=0.3', 'CNMC_3_blur_f=0.4',
                'CNMC_3_blur_f=0.5', 'CNMC_3_blur_f=0.6', 'CNMC_3_blur_f=0.7',
                'CNMC_3_blur_f=0.8', 'CNMC_3_blur_f=0.9', 'CNMC_3_blur_f=1.0',
                'CNMC_3_small_random_f=0.1', 'CNMC_3_small_random_f=0.2', 'CNMC_3_small_random_f=0.3', 'CNMC_3_small_random_f=0.4',
                'CNMC_3_small_random_f=0.5', 'CNMC_3_small_random_f=0.6', 'CNMC_3_small_random_f=0.7',
                'CNMC_3_small_random_f=0.8', 'CNMC_3_small_random_f=0.9', 'CNMC_3_small_random_f=1.0',
                'CNMC_3_small_easy_f=0.1', 'CNMC_3_small_easy_f=0.2', 'CNMC_3_small_easy_f=0.3', 'CNMC_3_small_easy_f=0.4',
                'CNMC_3_small_easy_f=0.5', 'CNMC_3_small_easy_f=0.6', 'CNMC_3_small_easy_f=0.7',
                'CNMC_3_small_easy_f=0.8', 'CNMC_3_small_easy_f=0.9', 'CNMC_3_small_easy_f=1.0',
                'CNMC_3_small_hard_f=0.1', 'CNMC_3_small_hard_f=0.2', 'CNMC_3_small_hard_f=0.3', 'CNMC_3_small_hard_f=0.4',
                'CNMC_3_small_hard_f=0.5', 'CNMC_3_small_hard_f=0.6', 'CNMC_3_small_hard_f=0.7',
                'CNMC_3_small_hard_f=0.8', 'CNMC_3_small_hard_f=0.9', 'CNMC_3_small_hard_f=1.0',
                'CNMC_3_small_clusters_f=0.1', 'CNMC_3_small_clusters_f=0.2', 'CNMC_3_small_clusters_f=0.3', 'CNMC_3_small_clusters_f=0.4',
                'CNMC_3_small_clusters_f=0.5', 'CNMC_3_small_clusters_f=0.6', 'CNMC_3_small_clusters_f=0.7',
                'CNMC_3_small_clusters_f=0.8', 'CNMC_3_small_clusters_f=0.9', 'CNMC_3_small_clusters_f=1.0',
                'CNMC_4', 'CNMC_4_image_rot_f=0.1', 'CNMC_4_image_rot_f=0.2',
                'CNMC_4_image_rot_f=0.3', 'CNMC_4_image_rot_f=0.4', 'CNMC_4_image_rot_f=0.5',
                'CNMC_4_image_rot_f=0.6', 'CNMC_4_image_rot_f=0.7', 'CNMC_4_image_rot_f=0.8',
                'CNMC_4_image_rot_f=0.9', 'CNMC_4_image_rot_f=1.0', 'CNMC_4_image_translation_f=0.1',
                'CNMC_4_image_translation_f=0.2', 'CNMC_4_image_translation_f=0.3', 'CNMC_4_image_translation_f=0.4',
                'CNMC_4_image_translation_f=0.5', 'CNMC_4_image_translation_f=0.6', 'CNMC_4_image_translation_f=0.7',
                'CNMC_4_image_translation_f=0.8', 'CNMC_4_image_translation_f=0.9', 'CNMC_4_image_translation_f=1.0',
                'CNMC_4_image_zoom_f=0.1', 'CNMC_4_image_zoom_f=0.2', 'CNMC_4_image_zoom_f=0.3',
                'CNMC_4_image_zoom_f=0.4', 'CNMC_4_image_zoom_f=0.5', 'CNMC_4_image_zoom_f=0.6',
                'CNMC_4_image_zoom_f=0.7', 'CNMC_4_image_zoom_f=0.8', 'CNMC_4_image_zoom_f=0.9',
                'CNMC_4_image_zoom_f=1.0', 'CNMC_4_add_noise_gaussian_f=0.1', 'CNMC_4_add_noise_gaussian_f=0.2',
                'CNMC_4_add_noise_gaussian_f=0.3', 'CNMC_4_add_noise_gaussian_f=0.4', 'CNMC_4_add_noise_gaussian_f=0.5',
                'CNMC_4_add_noise_gaussian_f=0.6', 'CNMC_4_add_noise_gaussian_f=0.7', 'CNMC_4_add_noise_gaussian_f=0.8',
                'CNMC_4_add_noise_gaussian_f=0.9', 'CNMC_4_add_noise_gaussian_f=1.0', 'CNMC_4_add_noise_poisson_f=0.1',
                'CNMC_4_add_noise_poisson_f=0.2', 'CNMC_4_add_noise_poisson_f=0.3', 'CNMC_4_add_noise_poisson_f=0.4',
                'CNMC_4_add_noise_poisson_f=0.5', 'CNMC_4_add_noise_poisson_f=0.6', 'CNMC_4_add_noise_poisson_f=0.7',
                'CNMC_4_add_noise_poisson_f=0.8', 'CNMC_4_add_noise_poisson_f=0.9', 'CNMC_4_add_noise_poisson_f=1.0',
                'CNMC_4_add_noise_salt_and_pepper_f=0.1', 'CNMC_4_add_noise_salt_and_pepper_f=0.2',
                'CNMC_4_add_noise_salt_and_pepper_f=0.3', 'CNMC_4_add_noise_salt_and_pepper_f=0.4',
                'CNMC_4_add_noise_salt_and_pepper_f=0.5', 'CNMC_4_add_noise_salt_and_pepper_f=0.6',
                'CNMC_4_add_noise_salt_and_pepper_f=0.7', 'CNMC_4_add_noise_salt_and_pepper_f=0.8',
                'CNMC_4_add_noise_salt_and_pepper_f=0.9', 'CNMC_4_add_noise_salt_and_pepper_f=1.0',
                'CNMC_4_add_noise_speckle_f=0.1', 'CNMC_4_add_noise_speckle_f=0.2', 'CNMC_4_add_noise_speckle_f=0.3',
                'CNMC_4_add_noise_speckle_f=0.4', 'CNMC_4_add_noise_speckle_f=0.5', 'CNMC_4_add_noise_speckle_f=0.6',
                'CNMC_4_add_noise_speckle_f=0.7', 'CNMC_4_add_noise_speckle_f=0.8', 'CNMC_4_add_noise_speckle_f=0.9',
                'CNMC_4_add_noise_speckle_f=1.0', 'CNMC_4_imbalance_classes_f=0.1', 'CNMC_4_imbalance_classes_f=0.2',
                'CNMC_4_imbalance_classes_f=0.3', 'CNMC_4_imbalance_classes_f=0.4', 'CNMC_4_imbalance_classes_f=0.5',
                'CNMC_4_imbalance_classes_f=0.6', 'CNMC_4_imbalance_classes_f=0.7', 'CNMC_4_imbalance_classes_f=0.8',
                'CNMC_4_imbalance_classes_f=0.9', 'CNMC_4_imbalance_classes_f=1.0', 'CNMC_4_grayscale_f=0.1',
                'CNMC_4_grayscale_f=0.2', 'CNMC_4_grayscale_f=0.3', 'CNMC_4_grayscale_f=0.4',
                'CNMC_4_grayscale_f=0.5', 'CNMC_4_grayscale_f=0.6', 'CNMC_4_grayscale_f=0.7',
                'CNMC_4_grayscale_f=0.8', 'CNMC_4_grayscale_f=0.9', 'CNMC_4_grayscale_f=1.0',
                'CNMC_4_hsv_f=0.1', 'CNMC_4_hsv_f=0.2', 'CNMC_4_hsv_f=0.3', 'CNMC_4_hsv_f=0.4',
                'CNMC_4_hsv_f=0.5', 'CNMC_4_hsv_f=0.6', 'CNMC_4_hsv_f=0.7',
                'CNMC_4_hsv_f=0.8', 'CNMC_4_hsv_f=0.9', 'CNMC_4_hsv_f=1.0',
                'CNMC_4_blur_f=0.1', 'CNMC_4_blur_f=0.2', 'CNMC_4_blur_f=0.3', 'CNMC_4_blur_f=0.4',
                'CNMC_4_blur_f=0.5', 'CNMC_4_blur_f=0.6', 'CNMC_4_blur_f=0.7',
                'CNMC_4_blur_f=0.8', 'CNMC_4_blur_f=0.9', 'CNMC_4_blur_f=1.0',
                'CNMC_4_small_random_f=0.1', 'CNMC_4_small_random_f=0.2', 'CNMC_4_small_random_f=0.3', 'CNMC_4_small_random_f=0.4',
                'CNMC_4_small_random_f=0.5', 'CNMC_4_small_random_f=0.6', 'CNMC_4_small_random_f=0.7',
                'CNMC_4_small_random_f=0.8', 'CNMC_4_small_random_f=0.9', 'CNMC_4_small_random_f=1.0',
                'CNMC_4_small_easy_f=0.1', 'CNMC_4_small_easy_f=0.2', 'CNMC_4_small_easy_f=0.3', 'CNMC_4_small_easy_f=0.4',
                'CNMC_4_small_easy_f=0.5', 'CNMC_4_small_easy_f=0.6', 'CNMC_4_small_easy_f=0.7',
                'CNMC_4_small_easy_f=0.8', 'CNMC_4_small_easy_f=0.9', 'CNMC_4_small_easy_f=1.0',
                'CNMC_4_small_hard_f=0.1', 'CNMC_4_small_hard_f=0.2', 'CNMC_4_small_hard_f=0.3', 'CNMC_4_small_hard_f=0.4',
                'CNMC_4_small_hard_f=0.5', 'CNMC_4_small_hard_f=0.6', 'CNMC_4_small_hard_f=0.7',
                'CNMC_4_small_hard_f=0.8', 'CNMC_4_small_hard_f=0.9', 'CNMC_4_small_hard_f=1.0',
                'CNMC_4_small_clusters_f=0.1', 'CNMC_4_small_clusters_f=0.2', 'CNMC_4_small_clusters_f=0.3', 'CNMC_4_small_clusters_f=0.4',
                'CNMC_4_small_clusters_f=0.5', 'CNMC_4_small_clusters_f=0.6', 'CNMC_4_small_clusters_f=0.7',
                'CNMC_4_small_clusters_f=0.8', 'CNMC_4_small_clusters_f=0.9', 'CNMC_4_small_clusters_f=1.0',
                'CNMC_5', 'CNMC_5_image_rot_f=0.1', 'CNMC_5_image_rot_f=0.2',
                'CNMC_5_image_rot_f=0.3', 'CNMC_5_image_rot_f=0.4', 'CNMC_5_image_rot_f=0.5',
                'CNMC_5_image_rot_f=0.6', 'CNMC_5_image_rot_f=0.7', 'CNMC_5_image_rot_f=0.8',
                'CNMC_5_image_rot_f=0.9', 'CNMC_5_image_rot_f=1.0', 'CNMC_5_image_translation_f=0.1',
                'CNMC_5_image_translation_f=0.2', 'CNMC_5_image_translation_f=0.3', 'CNMC_5_image_translation_f=0.4',
                'CNMC_5_image_translation_f=0.5', 'CNMC_5_image_translation_f=0.6', 'CNMC_5_image_translation_f=0.7',
                'CNMC_5_image_translation_f=0.8', 'CNMC_5_image_translation_f=0.9', 'CNMC_5_image_translation_f=1.0',
                'CNMC_5_image_zoom_f=0.1', 'CNMC_5_image_zoom_f=0.2', 'CNMC_5_image_zoom_f=0.3',
                'CNMC_5_image_zoom_f=0.4', 'CNMC_5_image_zoom_f=0.5', 'CNMC_5_image_zoom_f=0.6',
                'CNMC_5_image_zoom_f=0.7', 'CNMC_5_image_zoom_f=0.8', 'CNMC_5_image_zoom_f=0.9',
                'CNMC_5_image_zoom_f=1.0', 'CNMC_5_add_noise_gaussian_f=0.1', 'CNMC_5_add_noise_gaussian_f=0.2',
                'CNMC_5_add_noise_gaussian_f=0.3', 'CNMC_5_add_noise_gaussian_f=0.4', 'CNMC_5_add_noise_gaussian_f=0.5',
                'CNMC_5_add_noise_gaussian_f=0.6', 'CNMC_5_add_noise_gaussian_f=0.7', 'CNMC_5_add_noise_gaussian_f=0.8',
                'CNMC_5_add_noise_gaussian_f=0.9', 'CNMC_5_add_noise_gaussian_f=1.0', 'CNMC_5_add_noise_poisson_f=0.1',
                'CNMC_5_add_noise_poisson_f=0.2', 'CNMC_5_add_noise_poisson_f=0.3', 'CNMC_5_add_noise_poisson_f=0.4',
                'CNMC_5_add_noise_poisson_f=0.5', 'CNMC_5_add_noise_poisson_f=0.6', 'CNMC_5_add_noise_poisson_f=0.7',
                'CNMC_5_add_noise_poisson_f=0.8', 'CNMC_5_add_noise_poisson_f=0.9', 'CNMC_5_add_noise_poisson_f=1.0',
                'CNMC_5_add_noise_salt_and_pepper_f=0.1', 'CNMC_5_add_noise_salt_and_pepper_f=0.2',
                'CNMC_5_add_noise_salt_and_pepper_f=0.3', 'CNMC_5_add_noise_salt_and_pepper_f=0.4',
                'CNMC_5_add_noise_salt_and_pepper_f=0.5', 'CNMC_5_add_noise_salt_and_pepper_f=0.6',
                'CNMC_5_add_noise_salt_and_pepper_f=0.7', 'CNMC_5_add_noise_salt_and_pepper_f=0.8',
                'CNMC_5_add_noise_salt_and_pepper_f=0.9', 'CNMC_5_add_noise_salt_and_pepper_f=1.0',
                'CNMC_5_add_noise_speckle_f=0.1', 'CNMC_5_add_noise_speckle_f=0.2', 'CNMC_5_add_noise_speckle_f=0.3',
                'CNMC_5_add_noise_speckle_f=0.4', 'CNMC_5_add_noise_speckle_f=0.5', 'CNMC_5_add_noise_speckle_f=0.6',
                'CNMC_5_add_noise_speckle_f=0.7', 'CNMC_5_add_noise_speckle_f=0.8', 'CNMC_5_add_noise_speckle_f=0.9',
                'CNMC_5_add_noise_speckle_f=1.0', 'CNMC_5_imbalance_classes_f=0.1', 'CNMC_5_imbalance_classes_f=0.2',
                'CNMC_5_imbalance_classes_f=0.3', 'CNMC_5_imbalance_classes_f=0.4', 'CNMC_5_imbalance_classes_f=0.5',
                'CNMC_5_imbalance_classes_f=0.6', 'CNMC_5_imbalance_classes_f=0.7', 'CNMC_5_imbalance_classes_f=0.8',
                'CNMC_5_imbalance_classes_f=0.9', 'CNMC_5_imbalance_classes_f=1.0', 'CNMC_5_grayscale_f=0.1',
                'CNMC_5_grayscale_f=0.2', 'CNMC_5_grayscale_f=0.3', 'CNMC_5_grayscale_f=0.4',
                'CNMC_5_grayscale_f=0.5', 'CNMC_5_grayscale_f=0.6', 'CNMC_5_grayscale_f=0.7',
                'CNMC_5_grayscale_f=0.8', 'CNMC_5_grayscale_f=0.9', 'CNMC_5_grayscale_f=1.0',
                'CNMC_5_hsv_f=0.1', 'CNMC_5_hsv_f=0.2', 'CNMC_5_hsv_f=0.3', 'CNMC_5_hsv_f=0.4',
                'CNMC_5_hsv_f=0.5', 'CNMC_5_hsv_f=0.6', 'CNMC_5_hsv_f=0.7',
                'CNMC_5_hsv_f=0.8', 'CNMC_5_hsv_f=0.9', 'CNMC_5_hsv_f=1.0',
                'CNMC_5_blur_f=0.1', 'CNMC_5_blur_f=0.2', 'CNMC_5_blur_f=0.3', 'CNMC_5_blur_f=0.4',
                'CNMC_5_blur_f=0.5', 'CNMC_5_blur_f=0.6', 'CNMC_5_blur_f=0.7',
                'CNMC_5_blur_f=0.8', 'CNMC_5_blur_f=0.9', 'CNMC_5_blur_f=1.0',
                'CNMC_5_small_random_f=0.1', 'CNMC_5_small_random_f=0.2', 'CNMC_5_small_random_f=0.3', 'CNMC_5_small_random_f=0.4',
                'CNMC_5_small_random_f=0.5', 'CNMC_5_small_random_f=0.6', 'CNMC_5_small_random_f=0.7',
                'CNMC_5_small_random_f=0.8', 'CNMC_5_small_random_f=0.9', 'CNMC_5_small_random_f=1.0',
                'CNMC_5_small_easy_f=0.1', 'CNMC_5_small_easy_f=0.2', 'CNMC_5_small_easy_f=0.3', 'CNMC_5_small_easy_f=0.4',
                'CNMC_5_small_easy_f=0.5', 'CNMC_5_small_easy_f=0.6', 'CNMC_5_small_easy_f=0.7',
                'CNMC_5_small_easy_f=0.8', 'CNMC_5_small_easy_f=0.9', 'CNMC_5_small_easy_f=1.0',
                'CNMC_5_small_hard_f=0.1', 'CNMC_5_small_hard_f=0.2', 'CNMC_5_small_hard_f=0.3', 'CNMC_5_small_hard_f=0.4',
                'CNMC_5_small_hard_f=0.5', 'CNMC_5_small_hard_f=0.6', 'CNMC_5_small_hard_f=0.7',
                'CNMC_5_small_hard_f=0.8', 'CNMC_5_small_hard_f=0.9', 'CNMC_5_small_hard_f=1.0',
                'CNMC_5_small_clusters_f=0.1', 'CNMC_5_small_clusters_f=0.2', 'CNMC_5_small_clusters_f=0.3', 'CNMC_5_small_clusters_f=0.4',
                'CNMC_5_small_clusters_f=0.5', 'CNMC_5_small_clusters_f=0.6', 'CNMC_5_small_clusters_f=0.7',
                'CNMC_5_small_clusters_f=0.8', 'CNMC_5_small_clusters_f=0.9', 'CNMC_5_small_clusters_f=1.0',
                'CNMC_6', 'CNMC_6_image_rot_f=0.1', 'CNMC_6_image_rot_f=0.2',
                'CNMC_6_image_rot_f=0.3', 'CNMC_6_image_rot_f=0.4', 'CNMC_6_image_rot_f=0.5',
                'CNMC_6_image_rot_f=0.6', 'CNMC_6_image_rot_f=0.7', 'CNMC_6_image_rot_f=0.8',
                'CNMC_6_image_rot_f=0.9', 'CNMC_6_image_rot_f=1.0', 'CNMC_6_image_translation_f=0.1',
                'CNMC_6_image_translation_f=0.2', 'CNMC_6_image_translation_f=0.3', 'CNMC_6_image_translation_f=0.4',
                'CNMC_6_image_translation_f=0.5', 'CNMC_6_image_translation_f=0.6', 'CNMC_6_image_translation_f=0.7',
                'CNMC_6_image_translation_f=0.8', 'CNMC_6_image_translation_f=0.9', 'CNMC_6_image_translation_f=1.0',
                'CNMC_6_image_zoom_f=0.1', 'CNMC_6_image_zoom_f=0.2', 'CNMC_6_image_zoom_f=0.3',
                'CNMC_6_image_zoom_f=0.4', 'CNMC_6_image_zoom_f=0.5', 'CNMC_6_image_zoom_f=0.6',
                'CNMC_6_image_zoom_f=0.7', 'CNMC_6_image_zoom_f=0.8', 'CNMC_6_image_zoom_f=0.9',
                'CNMC_6_image_zoom_f=1.0', 'CNMC_6_add_noise_gaussian_f=0.1', 'CNMC_6_add_noise_gaussian_f=0.2',
                'CNMC_6_add_noise_gaussian_f=0.3', 'CNMC_6_add_noise_gaussian_f=0.4', 'CNMC_6_add_noise_gaussian_f=0.5',
                'CNMC_6_add_noise_gaussian_f=0.6', 'CNMC_6_add_noise_gaussian_f=0.7', 'CNMC_6_add_noise_gaussian_f=0.8',
                'CNMC_6_add_noise_gaussian_f=0.9', 'CNMC_6_add_noise_gaussian_f=1.0', 'CNMC_6_add_noise_poisson_f=0.1',
                'CNMC_6_add_noise_poisson_f=0.2', 'CNMC_6_add_noise_poisson_f=0.3', 'CNMC_6_add_noise_poisson_f=0.4',
                'CNMC_6_add_noise_poisson_f=0.5', 'CNMC_6_add_noise_poisson_f=0.6', 'CNMC_6_add_noise_poisson_f=0.7',
                'CNMC_6_add_noise_poisson_f=0.8', 'CNMC_6_add_noise_poisson_f=0.9', 'CNMC_6_add_noise_poisson_f=1.0',
                'CNMC_6_add_noise_salt_and_pepper_f=0.1', 'CNMC_6_add_noise_salt_and_pepper_f=0.2',
                'CNMC_6_add_noise_salt_and_pepper_f=0.3', 'CNMC_6_add_noise_salt_and_pepper_f=0.4',
                'CNMC_6_add_noise_salt_and_pepper_f=0.5', 'CNMC_6_add_noise_salt_and_pepper_f=0.6',
                'CNMC_6_add_noise_salt_and_pepper_f=0.7', 'CNMC_6_add_noise_salt_and_pepper_f=0.8',
                'CNMC_6_add_noise_salt_and_pepper_f=0.9', 'CNMC_6_add_noise_salt_and_pepper_f=1.0',
                'CNMC_6_add_noise_speckle_f=0.1', 'CNMC_6_add_noise_speckle_f=0.2', 'CNMC_6_add_noise_speckle_f=0.3',
                'CNMC_6_add_noise_speckle_f=0.4', 'CNMC_6_add_noise_speckle_f=0.5', 'CNMC_6_add_noise_speckle_f=0.6',
                'CNMC_6_add_noise_speckle_f=0.7', 'CNMC_6_add_noise_speckle_f=0.8', 'CNMC_6_add_noise_speckle_f=0.9',
                'CNMC_6_add_noise_speckle_f=1.0', 'CNMC_6_imbalance_classes_f=0.1', 'CNMC_6_imbalance_classes_f=0.2',
                'CNMC_6_imbalance_classes_f=0.3', 'CNMC_6_imbalance_classes_f=0.4', 'CNMC_6_imbalance_classes_f=0.5',
                'CNMC_6_imbalance_classes_f=0.6', 'CNMC_6_imbalance_classes_f=0.7', 'CNMC_6_imbalance_classes_f=0.8',
                'CNMC_6_imbalance_classes_f=0.9', 'CNMC_6_imbalance_classes_f=1.0', 'CNMC_6_grayscale_f=0.1',
                'CNMC_6_grayscale_f=0.2', 'CNMC_6_grayscale_f=0.3', 'CNMC_6_grayscale_f=0.4',
                'CNMC_6_grayscale_f=0.5', 'CNMC_6_grayscale_f=0.6', 'CNMC_6_grayscale_f=0.7',
                'CNMC_6_grayscale_f=0.8', 'CNMC_6_grayscale_f=0.9', 'CNMC_6_grayscale_f=1.0',
                'CNMC_6_hsv_f=0.1', 'CNMC_6_hsv_f=0.2', 'CNMC_6_hsv_f=0.3', 'CNMC_6_hsv_f=0.4',
                'CNMC_6_hsv_f=0.5', 'CNMC_6_hsv_f=0.6', 'CNMC_6_hsv_f=0.7',
                'CNMC_6_hsv_f=0.8', 'CNMC_6_hsv_f=0.9', 'CNMC_6_hsv_f=1.0',
                'CNMC_6_blur_f=0.1', 'CNMC_6_blur_f=0.2', 'CNMC_6_blur_f=0.3', 'CNMC_6_blur_f=0.4',
                'CNMC_6_blur_f=0.5', 'CNMC_6_blur_f=0.6', 'CNMC_6_blur_f=0.7',
                'CNMC_6_blur_f=0.8', 'CNMC_6_blur_f=0.9', 'CNMC_6_blur_f=1.0',
                'CNMC_6_small_random_f=0.1', 'CNMC_6_small_random_f=0.2', 'CNMC_6_small_random_f=0.3', 'CNMC_6_small_random_f=0.4',
                'CNMC_6_small_random_f=0.5', 'CNMC_6_small_random_f=0.6', 'CNMC_6_small_random_f=0.7',
                'CNMC_6_small_random_f=0.8', 'CNMC_6_small_random_f=0.9', 'CNMC_6_small_random_f=1.0',
                'CNMC_6_small_easy_f=0.1', 'CNMC_6_small_easy_f=0.2', 'CNMC_6_small_easy_f=0.3', 'CNMC_6_small_easy_f=0.4',
                'CNMC_6_small_easy_f=0.5', 'CNMC_6_small_easy_f=0.6', 'CNMC_6_small_easy_f=0.7',
                'CNMC_6_small_easy_f=0.8', 'CNMC_6_small_easy_f=0.9', 'CNMC_6_small_easy_f=1.0',
                'CNMC_6_small_hard_f=0.1', 'CNMC_6_small_hard_f=0.2', 'CNMC_6_small_hard_f=0.3', 'CNMC_6_small_hard_f=0.4',
                'CNMC_6_small_hard_f=0.5', 'CNMC_6_small_hard_f=0.6', 'CNMC_6_small_hard_f=0.7',
                'CNMC_6_small_hard_f=0.8', 'CNMC_6_small_hard_f=0.9', 'CNMC_6_small_hard_f=1.0',
                'CNMC_6_small_clusters_f=0.1', 'CNMC_6_small_clusters_f=0.2', 'CNMC_6_small_clusters_f=0.3', 'CNMC_6_small_clusters_f=0.4',
                'CNMC_6_small_clusters_f=0.5', 'CNMC_6_small_clusters_f=0.6', 'CNMC_6_small_clusters_f=0.7',
                'CNMC_6_small_clusters_f=0.8', 'CNMC_6_small_clusters_f=0.9', 'CNMC_6_small_clusters_f=1.0'],
        required=True,
        help='target dataset, name should be like [dataset]_[split]_[modification]_f=[fraction], e.g. "ISIC_2_grayscale_f=0.4"',
        metavar='target dataset, name should be like [dataset]_[split]_[modification]_f=[fraction], e.g. "ISIC_2_grayscale_f=0.4"')
    parser.add_argument('-s',
        '--source_dataset',
        choices=['ISIC_2', 'ISIC_2_image_rot_f=0.1', 'ISIC_2_image_rot_f=0.2',
                'ISIC_2_image_rot_f=0.3', 'ISIC_2_image_rot_f=0.4', 'ISIC_2_image_rot_f=0.5',
                'ISIC_2_image_rot_f=0.6', 'ISIC_2_image_rot_f=0.7', 'ISIC_2_image_rot_f=0.8',
                'ISIC_2_image_rot_f=0.9', 'ISIC_2_image_rot_f=1.0', 'ISIC_2_image_translation_f=0.1',
                'ISIC_2_image_translation_f=0.2', 'ISIC_2_image_translation_f=0.3', 'ISIC_2_image_translation_f=0.4',
                'ISIC_2_image_translation_f=0.5', 'ISIC_2_image_translation_f=0.6', 'ISIC_2_image_translation_f=0.7',
                'ISIC_2_image_translation_f=0.8', 'ISIC_2_image_translation_f=0.9', 'ISIC_2_image_translation_f=1.0',
                'ISIC_2_image_zoom_f=0.1', 'ISIC_2_image_zoom_f=0.2', 'ISIC_2_image_zoom_f=0.3',
                'ISIC_2_image_zoom_f=0.4', 'ISIC_2_image_zoom_f=0.5', 'ISIC_2_image_zoom_f=0.6',
                'ISIC_2_image_zoom_f=0.7', 'ISIC_2_image_zoom_f=0.8', 'ISIC_2_image_zoom_f=0.9',
                'ISIC_2_image_zoom_f=1.0', 'ISIC_2_add_noise_gaussian_f=0.1', 'ISIC_2_add_noise_gaussian_f=0.2',
                'ISIC_2_add_noise_gaussian_f=0.3', 'ISIC_2_add_noise_gaussian_f=0.4', 'ISIC_2_add_noise_gaussian_f=0.5',
                'ISIC_2_add_noise_gaussian_f=0.6', 'ISIC_2_add_noise_gaussian_f=0.7', 'ISIC_2_add_noise_gaussian_f=0.8',
                'ISIC_2_add_noise_gaussian_f=0.9', 'ISIC_2_add_noise_gaussian_f=1.0', 'ISIC_2_add_noise_poisson_f=0.1',
                'ISIC_2_add_noise_poisson_f=0.2', 'ISIC_2_add_noise_poisson_f=0.3', 'ISIC_2_add_noise_poisson_f=0.4',
                'ISIC_2_add_noise_poisson_f=0.5', 'ISIC_2_add_noise_poisson_f=0.6', 'ISIC_2_add_noise_poisson_f=0.7',
                'ISIC_2_add_noise_poisson_f=0.8', 'ISIC_2_add_noise_poisson_f=0.9', 'ISIC_2_add_noise_poisson_f=1.0',
                'ISIC_2_add_noise_salt_and_pepper_f=0.1', 'ISIC_2_add_noise_salt_and_pepper_f=0.2',
                'ISIC_2_add_noise_salt_and_pepper_f=0.3', 'ISIC_2_add_noise_salt_and_pepper_f=0.4',
                'ISIC_2_add_noise_salt_and_pepper_f=0.5', 'ISIC_2_add_noise_salt_and_pepper_f=0.6',
                'ISIC_2_add_noise_salt_and_pepper_f=0.7', 'ISIC_2_add_noise_salt_and_pepper_f=0.8',
                'ISIC_2_add_noise_salt_and_pepper_f=0.9', 'ISIC_2_add_noise_salt_and_pepper_f=1.0',
                'ISIC_2_add_noise_speckle_f=0.1', 'ISIC_2_add_noise_speckle_f=0.2', 'ISIC_2_add_noise_speckle_f=0.3',
                'ISIC_2_add_noise_speckle_f=0.4', 'ISIC_2_add_noise_speckle_f=0.5', 'ISIC_2_add_noise_speckle_f=0.6',
                'ISIC_2_add_noise_speckle_f=0.7', 'ISIC_2_add_noise_speckle_f=0.8', 'ISIC_2_add_noise_speckle_f=0.9',
                'ISIC_2_add_noise_speckle_f=1.0', 'ISIC_2_imbalance_classes_f=0.1', 'ISIC_2_imbalance_classes_f=0.2',
                'ISIC_2_imbalance_classes_f=0.3', 'ISIC_2_imbalance_classes_f=0.4', 'ISIC_2_imbalance_classes_f=0.5',
                'ISIC_2_imbalance_classes_f=0.6', 'ISIC_2_imbalance_classes_f=0.7', 'ISIC_2_imbalance_classes_f=0.8',
                'ISIC_2_imbalance_classes_f=0.9', 'ISIC_2_imbalance_classes_f=1.0', 'ISIC_2_grayscale_f=0.1',
                'ISIC_2_grayscale_f=0.2', 'ISIC_2_grayscale_f=0.3', 'ISIC_2_grayscale_f=0.4',
                'ISIC_2_grayscale_f=0.5', 'ISIC_2_grayscale_f=0.6', 'ISIC_2_grayscale_f=0.7',
                'ISIC_2_grayscale_f=0.8', 'ISIC_2_grayscale_f=0.9', 'ISIC_2_grayscale_f=1.0',
                'ISIC_2_hsv_f=0.1', 'ISIC_2_hsv_f=0.2', 'ISIC_2_hsv_f=0.3', 'ISIC_2_hsv_f=0.4',
                'ISIC_2_hsv_f=0.5', 'ISIC_2_hsv_f=0.6', 'ISIC_2_hsv_f=0.7',
                'ISIC_2_hsv_f=0.8', 'ISIC_2_hsv_f=0.9', 'ISIC_2_hsv_f=1.0',
                'ISIC_2_blur_f=0.1', 'ISIC_2_blur_f=0.2', 'ISIC_2_blur_f=0.3', 'ISIC_2_blur_f=0.4',
                'ISIC_2_blur_f=0.5', 'ISIC_2_blur_f=0.6', 'ISIC_2_blur_f=0.7',
                'ISIC_2_blur_f=0.8', 'ISIC_2_blur_f=0.9', 'ISIC_2_blur_f=1.0',
                'ISIC_2_small_random_f=0.1', 'ISIC_2_small_random_f=0.2', 'ISIC_2_small_random_f=0.3', 'ISIC_2_small_random_f=0.4',
                'ISIC_2_small_random_f=0.5', 'ISIC_2_small_random_f=0.6', 'ISIC_2_small_random_f=0.7',
                'ISIC_2_small_random_f=0.8', 'ISIC_2_small_random_f=0.9', 'ISIC_2_small_random_f=1.0',
                'ISIC_2_small_easy_f=0.1', 'ISIC_2_small_easy_f=0.2', 'ISIC_2_small_easy_f=0.3', 'ISIC_2_small_easy_f=0.4',
                'ISIC_2_small_easy_f=0.5', 'ISIC_2_small_easy_f=0.6', 'ISIC_2_small_easy_f=0.7',
                'ISIC_2_small_easy_f=0.8', 'ISIC_2_small_easy_f=0.9', 'ISIC_2_small_easy_f=1.0',
                'ISIC_2_small_hard_f=0.1', 'ISIC_2_small_hard_f=0.2', 'ISIC_2_small_hard_f=0.3', 'ISIC_2_small_hard_f=0.4',
                'ISIC_2_small_hard_f=0.5', 'ISIC_2_small_hard_f=0.6', 'ISIC_2_small_hard_f=0.7',
                'ISIC_2_small_hard_f=0.8', 'ISIC_2_small_hard_f=0.9', 'ISIC_2_small_hard_f=1.0',
                'ISIC_2_small_clusters_f=0.1', 'ISIC_2_small_clusters_f=0.2', 'ISIC_2_small_clusters_f=0.3', 'ISIC_2_small_clusters_f=0.4',
                'ISIC_2_small_clusters_f=0.5', 'ISIC_2_small_clusters_f=0.6', 'ISIC_2_small_clusters_f=0.7',
                'ISIC_2_small_clusters_f=0.8', 'ISIC_2_small_clusters_f=0.9', 'ISIC_2_small_clusters_f=1.0',
                'ISIC_3', 'ISIC_3_image_rot_f=0.1', 'ISIC_3_image_rot_f=0.2',
                'ISIC_3_image_rot_f=0.3', 'ISIC_3_image_rot_f=0.4', 'ISIC_3_image_rot_f=0.5',
                'ISIC_3_image_rot_f=0.6', 'ISIC_3_image_rot_f=0.7', 'ISIC_3_image_rot_f=0.8',
                'ISIC_3_image_rot_f=0.9', 'ISIC_3_image_rot_f=1.0', 'ISIC_3_image_translation_f=0.1',
                'ISIC_3_image_translation_f=0.2', 'ISIC_3_image_translation_f=0.3', 'ISIC_3_image_translation_f=0.4',
                'ISIC_3_image_translation_f=0.5', 'ISIC_3_image_translation_f=0.6', 'ISIC_3_image_translation_f=0.7',
                'ISIC_3_image_translation_f=0.8', 'ISIC_3_image_translation_f=0.9', 'ISIC_3_image_translation_f=1.0',
                'ISIC_3_image_zoom_f=0.1', 'ISIC_3_image_zoom_f=0.2', 'ISIC_3_image_zoom_f=0.3',
                'ISIC_3_image_zoom_f=0.4', 'ISIC_3_image_zoom_f=0.5', 'ISIC_3_image_zoom_f=0.6',
                'ISIC_3_image_zoom_f=0.7', 'ISIC_3_image_zoom_f=0.8', 'ISIC_3_image_zoom_f=0.9',
                'ISIC_3_image_zoom_f=1.0', 'ISIC_3_add_noise_gaussian_f=0.1', 'ISIC_3_add_noise_gaussian_f=0.2',
                'ISIC_3_add_noise_gaussian_f=0.3', 'ISIC_3_add_noise_gaussian_f=0.4', 'ISIC_3_add_noise_gaussian_f=0.5',
                'ISIC_3_add_noise_gaussian_f=0.6', 'ISIC_3_add_noise_gaussian_f=0.7', 'ISIC_3_add_noise_gaussian_f=0.8',
                'ISIC_3_add_noise_gaussian_f=0.9', 'ISIC_3_add_noise_gaussian_f=1.0', 'ISIC_3_add_noise_poisson_f=0.1',
                'ISIC_3_add_noise_poisson_f=0.2', 'ISIC_3_add_noise_poisson_f=0.3', 'ISIC_3_add_noise_poisson_f=0.4',
                'ISIC_3_add_noise_poisson_f=0.5', 'ISIC_3_add_noise_poisson_f=0.6', 'ISIC_3_add_noise_poisson_f=0.7',
                'ISIC_3_add_noise_poisson_f=0.8', 'ISIC_3_add_noise_poisson_f=0.9', 'ISIC_3_add_noise_poisson_f=1.0',
                'ISIC_3_add_noise_salt_and_pepper_f=0.1', 'ISIC_3_add_noise_salt_and_pepper_f=0.2',
                'ISIC_3_add_noise_salt_and_pepper_f=0.3', 'ISIC_3_add_noise_salt_and_pepper_f=0.4',
                'ISIC_3_add_noise_salt_and_pepper_f=0.5', 'ISIC_3_add_noise_salt_and_pepper_f=0.6',
                'ISIC_3_add_noise_salt_and_pepper_f=0.7', 'ISIC_3_add_noise_salt_and_pepper_f=0.8',
                'ISIC_3_add_noise_salt_and_pepper_f=0.9', 'ISIC_3_add_noise_salt_and_pepper_f=1.0',
                'ISIC_3_add_noise_speckle_f=0.1', 'ISIC_3_add_noise_speckle_f=0.2', 'ISIC_3_add_noise_speckle_f=0.3',
                'ISIC_3_add_noise_speckle_f=0.4', 'ISIC_3_add_noise_speckle_f=0.5', 'ISIC_3_add_noise_speckle_f=0.6',
                'ISIC_3_add_noise_speckle_f=0.7', 'ISIC_3_add_noise_speckle_f=0.8', 'ISIC_3_add_noise_speckle_f=0.9',
                'ISIC_3_add_noise_speckle_f=1.0', 'ISIC_3_imbalance_classes_f=0.1', 'ISIC_3_imbalance_classes_f=0.2',
                'ISIC_3_imbalance_classes_f=0.3', 'ISIC_3_imbalance_classes_f=0.4', 'ISIC_3_imbalance_classes_f=0.5',
                'ISIC_3_imbalance_classes_f=0.6', 'ISIC_3_imbalance_classes_f=0.7', 'ISIC_3_imbalance_classes_f=0.8',
                'ISIC_3_imbalance_classes_f=0.9', 'ISIC_3_imbalance_classes_f=1.0', 'ISIC_3_grayscale_f=0.1',
                'ISIC_3_grayscale_f=0.2', 'ISIC_3_grayscale_f=0.3', 'ISIC_3_grayscale_f=0.4',
                'ISIC_3_grayscale_f=0.5', 'ISIC_3_grayscale_f=0.6', 'ISIC_3_grayscale_f=0.7',
                'ISIC_3_grayscale_f=0.8', 'ISIC_3_grayscale_f=0.9', 'ISIC_3_grayscale_f=1.0',
                'ISIC_3_hsv_f=0.1', 'ISIC_3_hsv_f=0.2', 'ISIC_3_hsv_f=0.3', 'ISIC_3_hsv_f=0.4',
                'ISIC_3_hsv_f=0.5', 'ISIC_3_hsv_f=0.6', 'ISIC_3_hsv_f=0.7',
                'ISIC_3_hsv_f=0.8', 'ISIC_3_hsv_f=0.9', 'ISIC_3_hsv_f=1.0',
                'ISIC_3_blur_f=0.1', 'ISIC_3_blur_f=0.2', 'ISIC_3_blur_f=0.3', 'ISIC_3_blur_f=0.4',
                'ISIC_3_blur_f=0.5', 'ISIC_3_blur_f=0.6', 'ISIC_3_blur_f=0.7',
                'ISIC_3_blur_f=0.8', 'ISIC_3_blur_f=0.9', 'ISIC_3_blur_f=1.0',
                'ISIC_3_small_random_f=0.1', 'ISIC_3_small_random_f=0.2', 'ISIC_3_small_random_f=0.3', 'ISIC_3_small_random_f=0.4',
                'ISIC_3_small_random_f=0.5', 'ISIC_3_small_random_f=0.6', 'ISIC_3_small_random_f=0.7',
                'ISIC_3_small_random_f=0.8', 'ISIC_3_small_random_f=0.9', 'ISIC_3_small_random_f=1.0',
                'ISIC_3_small_easy_f=0.1', 'ISIC_3_small_easy_f=0.2', 'ISIC_3_small_easy_f=0.3', 'ISIC_3_small_easy_f=0.4',
                'ISIC_3_small_easy_f=0.5', 'ISIC_3_small_easy_f=0.6', 'ISIC_3_small_easy_f=0.7',
                'ISIC_3_small_easy_f=0.8', 'ISIC_3_small_easy_f=0.9', 'ISIC_3_small_easy_f=1.0',
                'ISIC_3_small_hard_f=0.1', 'ISIC_3_small_hard_f=0.2', 'ISIC_3_small_hard_f=0.3', 'ISIC_3_small_hard_f=0.4',
                'ISIC_3_small_hard_f=0.5', 'ISIC_3_small_hard_f=0.6', 'ISIC_3_small_hard_f=0.7',
                'ISIC_3_small_hard_f=0.8', 'ISIC_3_small_hard_f=0.9', 'ISIC_3_small_hard_f=1.0',
                'ISIC_3_small_clusters_f=0.1', 'ISIC_3_small_clusters_f=0.2', 'ISIC_3_small_clusters_f=0.3', 'ISIC_3_small_clusters_f=0.4',
                'ISIC_3_small_clusters_f=0.5', 'ISIC_3_small_clusters_f=0.6', 'ISIC_3_small_clusters_f=0.7',
                'ISIC_3_small_clusters_f=0.8', 'ISIC_3_small_clusters_f=0.9', 'ISIC_3_small_clusters_f=1.0',
                'ISIC_4', 'ISIC_4_image_rot_f=0.1', 'ISIC_4_image_rot_f=0.2',
                'ISIC_4_image_rot_f=0.3', 'ISIC_4_image_rot_f=0.4', 'ISIC_4_image_rot_f=0.5',
                'ISIC_4_image_rot_f=0.6', 'ISIC_4_image_rot_f=0.7', 'ISIC_4_image_rot_f=0.8',
                'ISIC_4_image_rot_f=0.9', 'ISIC_4_image_rot_f=1.0', 'ISIC_4_image_translation_f=0.1',
                'ISIC_4_image_translation_f=0.2', 'ISIC_4_image_translation_f=0.3', 'ISIC_4_image_translation_f=0.4',
                'ISIC_4_image_translation_f=0.5', 'ISIC_4_image_translation_f=0.6', 'ISIC_4_image_translation_f=0.7',
                'ISIC_4_image_translation_f=0.8', 'ISIC_4_image_translation_f=0.9', 'ISIC_4_image_translation_f=1.0',
                'ISIC_4_image_zoom_f=0.1', 'ISIC_4_image_zoom_f=0.2', 'ISIC_4_image_zoom_f=0.3',
                'ISIC_4_image_zoom_f=0.4', 'ISIC_4_image_zoom_f=0.5', 'ISIC_4_image_zoom_f=0.6',
                'ISIC_4_image_zoom_f=0.7', 'ISIC_4_image_zoom_f=0.8', 'ISIC_4_image_zoom_f=0.9',
                'ISIC_4_image_zoom_f=1.0', 'ISIC_4_add_noise_gaussian_f=0.1', 'ISIC_4_add_noise_gaussian_f=0.2',
                'ISIC_4_add_noise_gaussian_f=0.3', 'ISIC_4_add_noise_gaussian_f=0.4', 'ISIC_4_add_noise_gaussian_f=0.5',
                'ISIC_4_add_noise_gaussian_f=0.6', 'ISIC_4_add_noise_gaussian_f=0.7', 'ISIC_4_add_noise_gaussian_f=0.8',
                'ISIC_4_add_noise_gaussian_f=0.9', 'ISIC_4_add_noise_gaussian_f=1.0', 'ISIC_4_add_noise_poisson_f=0.1',
                'ISIC_4_add_noise_poisson_f=0.2', 'ISIC_4_add_noise_poisson_f=0.3', 'ISIC_4_add_noise_poisson_f=0.4',
                'ISIC_4_add_noise_poisson_f=0.5', 'ISIC_4_add_noise_poisson_f=0.6', 'ISIC_4_add_noise_poisson_f=0.7',
                'ISIC_4_add_noise_poisson_f=0.8', 'ISIC_4_add_noise_poisson_f=0.9', 'ISIC_4_add_noise_poisson_f=1.0',
                'ISIC_4_add_noise_salt_and_pepper_f=0.1', 'ISIC_4_add_noise_salt_and_pepper_f=0.2',
                'ISIC_4_add_noise_salt_and_pepper_f=0.3', 'ISIC_4_add_noise_salt_and_pepper_f=0.4',
                'ISIC_4_add_noise_salt_and_pepper_f=0.5', 'ISIC_4_add_noise_salt_and_pepper_f=0.6',
                'ISIC_4_add_noise_salt_and_pepper_f=0.7', 'ISIC_4_add_noise_salt_and_pepper_f=0.8',
                'ISIC_4_add_noise_salt_and_pepper_f=0.9', 'ISIC_4_add_noise_salt_and_pepper_f=1.0',
                'ISIC_4_add_noise_speckle_f=0.1', 'ISIC_4_add_noise_speckle_f=0.2', 'ISIC_4_add_noise_speckle_f=0.3',
                'ISIC_4_add_noise_speckle_f=0.4', 'ISIC_4_add_noise_speckle_f=0.5', 'ISIC_4_add_noise_speckle_f=0.6',
                'ISIC_4_add_noise_speckle_f=0.7', 'ISIC_4_add_noise_speckle_f=0.8', 'ISIC_4_add_noise_speckle_f=0.9',
                'ISIC_4_add_noise_speckle_f=1.0', 'ISIC_4_imbalance_classes_f=0.1', 'ISIC_4_imbalance_classes_f=0.2',
                'ISIC_4_imbalance_classes_f=0.3', 'ISIC_4_imbalance_classes_f=0.4', 'ISIC_4_imbalance_classes_f=0.5',
                'ISIC_4_imbalance_classes_f=0.6', 'ISIC_4_imbalance_classes_f=0.7', 'ISIC_4_imbalance_classes_f=0.8',
                'ISIC_4_imbalance_classes_f=0.9', 'ISIC_4_imbalance_classes_f=1.0', 'ISIC_4_grayscale_f=0.1',
                'ISIC_4_grayscale_f=0.2', 'ISIC_4_grayscale_f=0.3', 'ISIC_4_grayscale_f=0.4',
                'ISIC_4_grayscale_f=0.5', 'ISIC_4_grayscale_f=0.6', 'ISIC_4_grayscale_f=0.7',
                'ISIC_4_grayscale_f=0.8', 'ISIC_4_grayscale_f=0.9', 'ISIC_4_grayscale_f=1.0',
                'ISIC_4_hsv_f=0.1', 'ISIC_4_hsv_f=0.2', 'ISIC_4_hsv_f=0.3', 'ISIC_4_hsv_f=0.4',
                'ISIC_4_hsv_f=0.5', 'ISIC_4_hsv_f=0.6', 'ISIC_4_hsv_f=0.7',
                'ISIC_4_hsv_f=0.8', 'ISIC_4_hsv_f=0.9', 'ISIC_4_hsv_f=1.0',
                'ISIC_4_blur_f=0.1', 'ISIC_4_blur_f=0.2', 'ISIC_4_blur_f=0.3', 'ISIC_4_blur_f=0.4',
                'ISIC_4_blur_f=0.5', 'ISIC_4_blur_f=0.6', 'ISIC_4_blur_f=0.7',
                'ISIC_4_blur_f=0.8', 'ISIC_4_blur_f=0.9', 'ISIC_4_blur_f=1.0',
                'ISIC_4_small_random_f=0.1', 'ISIC_4_small_random_f=0.2', 'ISIC_4_small_random_f=0.3', 'ISIC_4_small_random_f=0.4',
                'ISIC_4_small_random_f=0.5', 'ISIC_4_small_random_f=0.6', 'ISIC_4_small_random_f=0.7',
                'ISIC_4_small_random_f=0.8', 'ISIC_4_small_random_f=0.9', 'ISIC_4_small_random_f=1.0',
                'ISIC_4_small_easy_f=0.1', 'ISIC_4_small_easy_f=0.2', 'ISIC_4_small_easy_f=0.3', 'ISIC_4_small_easy_f=0.4',
                'ISIC_4_small_easy_f=0.5', 'ISIC_4_small_easy_f=0.6', 'ISIC_4_small_easy_f=0.7',
                'ISIC_4_small_easy_f=0.8', 'ISIC_4_small_easy_f=0.9', 'ISIC_4_small_easy_f=1.0',
                'ISIC_4_small_hard_f=0.1', 'ISIC_4_small_hard_f=0.2', 'ISIC_4_small_hard_f=0.3', 'ISIC_4_small_hard_f=0.4',
                'ISIC_4_small_hard_f=0.5', 'ISIC_4_small_hard_f=0.6', 'ISIC_4_small_hard_f=0.7',
                'ISIC_4_small_hard_f=0.8', 'ISIC_4_small_hard_f=0.9', 'ISIC_4_small_hard_f=1.0',
                'ISIC_4_small_clusters_f=0.1', 'ISIC_4_small_clusters_f=0.2', 'ISIC_4_small_clusters_f=0.3', 'ISIC_4_small_clusters_f=0.4',
                'ISIC_4_small_clusters_f=0.5', 'ISIC_4_small_clusters_f=0.6', 'ISIC_4_small_clusters_f=0.7',
                'ISIC_4_small_clusters_f=0.8', 'ISIC_4_small_clusters_f=0.9', 'ISIC_4_small_clusters_f=1.0',
                'ISIC_5', 'ISIC_5_image_rot_f=0.1', 'ISIC_5_image_rot_f=0.2',
                'ISIC_5_image_rot_f=0.3', 'ISIC_5_image_rot_f=0.4', 'ISIC_5_image_rot_f=0.5',
                'ISIC_5_image_rot_f=0.6', 'ISIC_5_image_rot_f=0.7', 'ISIC_5_image_rot_f=0.8',
                'ISIC_5_image_rot_f=0.9', 'ISIC_5_image_rot_f=1.0', 'ISIC_5_image_translation_f=0.1',
                'ISIC_5_image_translation_f=0.2', 'ISIC_5_image_translation_f=0.3', 'ISIC_5_image_translation_f=0.4',
                'ISIC_5_image_translation_f=0.5', 'ISIC_5_image_translation_f=0.6', 'ISIC_5_image_translation_f=0.7',
                'ISIC_5_image_translation_f=0.8', 'ISIC_5_image_translation_f=0.9', 'ISIC_5_image_translation_f=1.0',
                'ISIC_5_image_zoom_f=0.1', 'ISIC_5_image_zoom_f=0.2', 'ISIC_5_image_zoom_f=0.3',
                'ISIC_5_image_zoom_f=0.4', 'ISIC_5_image_zoom_f=0.5', 'ISIC_5_image_zoom_f=0.6',
                'ISIC_5_image_zoom_f=0.7', 'ISIC_5_image_zoom_f=0.8', 'ISIC_5_image_zoom_f=0.9',
                'ISIC_5_image_zoom_f=1.0', 'ISIC_5_add_noise_gaussian_f=0.1', 'ISIC_5_add_noise_gaussian_f=0.2',
                'ISIC_5_add_noise_gaussian_f=0.3', 'ISIC_5_add_noise_gaussian_f=0.4', 'ISIC_5_add_noise_gaussian_f=0.5',
                'ISIC_5_add_noise_gaussian_f=0.6', 'ISIC_5_add_noise_gaussian_f=0.7', 'ISIC_5_add_noise_gaussian_f=0.8',
                'ISIC_5_add_noise_gaussian_f=0.9', 'ISIC_5_add_noise_gaussian_f=1.0', 'ISIC_5_add_noise_poisson_f=0.1',
                'ISIC_5_add_noise_poisson_f=0.2', 'ISIC_5_add_noise_poisson_f=0.3', 'ISIC_5_add_noise_poisson_f=0.4',
                'ISIC_5_add_noise_poisson_f=0.5', 'ISIC_5_add_noise_poisson_f=0.6', 'ISIC_5_add_noise_poisson_f=0.7',
                'ISIC_5_add_noise_poisson_f=0.8', 'ISIC_5_add_noise_poisson_f=0.9', 'ISIC_5_add_noise_poisson_f=1.0',
                'ISIC_5_add_noise_salt_and_pepper_f=0.1', 'ISIC_5_add_noise_salt_and_pepper_f=0.2',
                'ISIC_5_add_noise_salt_and_pepper_f=0.3', 'ISIC_5_add_noise_salt_and_pepper_f=0.4',
                'ISIC_5_add_noise_salt_and_pepper_f=0.5', 'ISIC_5_add_noise_salt_and_pepper_f=0.6',
                'ISIC_5_add_noise_salt_and_pepper_f=0.7', 'ISIC_5_add_noise_salt_and_pepper_f=0.8',
                'ISIC_5_add_noise_salt_and_pepper_f=0.9', 'ISIC_5_add_noise_salt_and_pepper_f=1.0',
                'ISIC_5_add_noise_speckle_f=0.1', 'ISIC_5_add_noise_speckle_f=0.2', 'ISIC_5_add_noise_speckle_f=0.3',
                'ISIC_5_add_noise_speckle_f=0.4', 'ISIC_5_add_noise_speckle_f=0.5', 'ISIC_5_add_noise_speckle_f=0.6',
                'ISIC_5_add_noise_speckle_f=0.7', 'ISIC_5_add_noise_speckle_f=0.8', 'ISIC_5_add_noise_speckle_f=0.9',
                'ISIC_5_add_noise_speckle_f=1.0', 'ISIC_5_imbalance_classes_f=0.1', 'ISIC_5_imbalance_classes_f=0.2',
                'ISIC_5_imbalance_classes_f=0.3', 'ISIC_5_imbalance_classes_f=0.4', 'ISIC_5_imbalance_classes_f=0.5',
                'ISIC_5_imbalance_classes_f=0.6', 'ISIC_5_imbalance_classes_f=0.7', 'ISIC_5_imbalance_classes_f=0.8',
                'ISIC_5_imbalance_classes_f=0.9', 'ISIC_5_imbalance_classes_f=1.0', 'ISIC_5_grayscale_f=0.1',
                'ISIC_5_grayscale_f=0.2', 'ISIC_5_grayscale_f=0.3', 'ISIC_5_grayscale_f=0.4',
                'ISIC_5_grayscale_f=0.5', 'ISIC_5_grayscale_f=0.6', 'ISIC_5_grayscale_f=0.7',
                'ISIC_5_grayscale_f=0.8', 'ISIC_5_grayscale_f=0.9', 'ISIC_5_grayscale_f=1.0',
                'ISIC_5_hsv_f=0.1', 'ISIC_5_hsv_f=0.2', 'ISIC_5_hsv_f=0.3', 'ISIC_5_hsv_f=0.4',
                'ISIC_5_hsv_f=0.5', 'ISIC_5_hsv_f=0.6', 'ISIC_5_hsv_f=0.7',
                'ISIC_5_hsv_f=0.8', 'ISIC_5_hsv_f=0.9', 'ISIC_5_hsv_f=1.0',
                'ISIC_5_blur_f=0.1', 'ISIC_5_blur_f=0.2', 'ISIC_5_blur_f=0.3', 'ISIC_5_blur_f=0.4',
                'ISIC_5_blur_f=0.5', 'ISIC_5_blur_f=0.6', 'ISIC_5_blur_f=0.7',
                'ISIC_5_blur_f=0.8', 'ISIC_5_blur_f=0.9', 'ISIC_5_blur_f=1.0',
                'ISIC_5_small_random_f=0.1', 'ISIC_5_small_random_f=0.2', 'ISIC_5_small_random_f=0.3', 'ISIC_5_small_random_f=0.4',
                'ISIC_5_small_random_f=0.5', 'ISIC_5_small_random_f=0.6', 'ISIC_5_small_random_f=0.7',
                'ISIC_5_small_random_f=0.8', 'ISIC_5_small_random_f=0.9', 'ISIC_5_small_random_f=1.0',
                'ISIC_5_small_easy_f=0.1', 'ISIC_5_small_easy_f=0.2', 'ISIC_5_small_easy_f=0.3', 'ISIC_5_small_easy_f=0.4',
                'ISIC_5_small_easy_f=0.5', 'ISIC_5_small_easy_f=0.6', 'ISIC_5_small_easy_f=0.7',
                'ISIC_5_small_easy_f=0.8', 'ISIC_5_small_easy_f=0.9', 'ISIC_5_small_easy_f=1.0',
                'ISIC_5_small_hard_f=0.1', 'ISIC_5_small_hard_f=0.2', 'ISIC_5_small_hard_f=0.3', 'ISIC_5_small_hard_f=0.4',
                'ISIC_5_small_hard_f=0.5', 'ISIC_5_small_hard_f=0.6', 'ISIC_5_small_hard_f=0.7',
                'ISIC_5_small_hard_f=0.8', 'ISIC_5_small_hard_f=0.9', 'ISIC_5_small_hard_f=1.0',
                'ISIC_5_small_clusters_f=0.1', 'ISIC_5_small_clusters_f=0.2', 'ISIC_5_small_clusters_f=0.3', 'ISIC_5_small_clusters_f=0.4',
                'ISIC_5_small_clusters_f=0.5', 'ISIC_5_small_clusters_f=0.6', 'ISIC_5_small_clusters_f=0.7',
                'ISIC_5_small_clusters_f=0.8', 'ISIC_5_small_clusters_f=0.9', 'ISIC_5_small_clusters_f=1.0',
                'ISIC_6', 'ISIC_6_image_rot_f=0.1', 'ISIC_6_image_rot_f=0.2',
                'ISIC_6_image_rot_f=0.3', 'ISIC_6_image_rot_f=0.4', 'ISIC_6_image_rot_f=0.5',
                'ISIC_6_image_rot_f=0.6', 'ISIC_6_image_rot_f=0.7', 'ISIC_6_image_rot_f=0.8',
                'ISIC_6_image_rot_f=0.9', 'ISIC_6_image_rot_f=1.0', 'ISIC_6_image_translation_f=0.1',
                'ISIC_6_image_translation_f=0.2', 'ISIC_6_image_translation_f=0.3', 'ISIC_6_image_translation_f=0.4',
                'ISIC_6_image_translation_f=0.5', 'ISIC_6_image_translation_f=0.6', 'ISIC_6_image_translation_f=0.7',
                'ISIC_6_image_translation_f=0.8', 'ISIC_6_image_translation_f=0.9', 'ISIC_6_image_translation_f=1.0',
                'ISIC_6_image_zoom_f=0.1', 'ISIC_6_image_zoom_f=0.2', 'ISIC_6_image_zoom_f=0.3',
                'ISIC_6_image_zoom_f=0.4', 'ISIC_6_image_zoom_f=0.5', 'ISIC_6_image_zoom_f=0.6',
                'ISIC_6_image_zoom_f=0.7', 'ISIC_6_image_zoom_f=0.8', 'ISIC_6_image_zoom_f=0.9',
                'ISIC_6_image_zoom_f=1.0', 'ISIC_6_add_noise_gaussian_f=0.1', 'ISIC_6_add_noise_gaussian_f=0.2',
                'ISIC_6_add_noise_gaussian_f=0.3', 'ISIC_6_add_noise_gaussian_f=0.4', 'ISIC_6_add_noise_gaussian_f=0.5',
                'ISIC_6_add_noise_gaussian_f=0.6', 'ISIC_6_add_noise_gaussian_f=0.7', 'ISIC_6_add_noise_gaussian_f=0.8',
                'ISIC_6_add_noise_gaussian_f=0.9', 'ISIC_6_add_noise_gaussian_f=1.0', 'ISIC_6_add_noise_poisson_f=0.1',
                'ISIC_6_add_noise_poisson_f=0.2', 'ISIC_6_add_noise_poisson_f=0.3', 'ISIC_6_add_noise_poisson_f=0.4',
                'ISIC_6_add_noise_poisson_f=0.5', 'ISIC_6_add_noise_poisson_f=0.6', 'ISIC_6_add_noise_poisson_f=0.7',
                'ISIC_6_add_noise_poisson_f=0.8', 'ISIC_6_add_noise_poisson_f=0.9', 'ISIC_6_add_noise_poisson_f=1.0',
                'ISIC_6_add_noise_salt_and_pepper_f=0.1', 'ISIC_6_add_noise_salt_and_pepper_f=0.2',
                'ISIC_6_add_noise_salt_and_pepper_f=0.3', 'ISIC_6_add_noise_salt_and_pepper_f=0.4',
                'ISIC_6_add_noise_salt_and_pepper_f=0.5', 'ISIC_6_add_noise_salt_and_pepper_f=0.6',
                'ISIC_6_add_noise_salt_and_pepper_f=0.7', 'ISIC_6_add_noise_salt_and_pepper_f=0.8',
                'ISIC_6_add_noise_salt_and_pepper_f=0.9', 'ISIC_6_add_noise_salt_and_pepper_f=1.0',
                'ISIC_6_add_noise_speckle_f=0.1', 'ISIC_6_add_noise_speckle_f=0.2', 'ISIC_6_add_noise_speckle_f=0.3',
                'ISIC_6_add_noise_speckle_f=0.4', 'ISIC_6_add_noise_speckle_f=0.5', 'ISIC_6_add_noise_speckle_f=0.6',
                'ISIC_6_add_noise_speckle_f=0.7', 'ISIC_6_add_noise_speckle_f=0.8', 'ISIC_6_add_noise_speckle_f=0.9',
                'ISIC_6_add_noise_speckle_f=1.0', 'ISIC_6_imbalance_classes_f=0.1', 'ISIC_6_imbalance_classes_f=0.2',
                'ISIC_6_imbalance_classes_f=0.3', 'ISIC_6_imbalance_classes_f=0.4', 'ISIC_6_imbalance_classes_f=0.5',
                'ISIC_6_imbalance_classes_f=0.6', 'ISIC_6_imbalance_classes_f=0.7', 'ISIC_6_imbalance_classes_f=0.8',
                'ISIC_6_imbalance_classes_f=0.9', 'ISIC_6_imbalance_classes_f=1.0', 'ISIC_6_grayscale_f=0.1',
                'ISIC_6_grayscale_f=0.2', 'ISIC_6_grayscale_f=0.3', 'ISIC_6_grayscale_f=0.4',
                'ISIC_6_grayscale_f=0.5', 'ISIC_6_grayscale_f=0.6', 'ISIC_6_grayscale_f=0.7',
                'ISIC_6_grayscale_f=0.8', 'ISIC_6_grayscale_f=0.9', 'ISIC_6_grayscale_f=1.0',
                'ISIC_6_hsv_f=0.1', 'ISIC_6_hsv_f=0.2', 'ISIC_6_hsv_f=0.3', 'ISIC_6_hsv_f=0.4',
                'ISIC_6_hsv_f=0.5', 'ISIC_6_hsv_f=0.6', 'ISIC_6_hsv_f=0.7',
                'ISIC_6_hsv_f=0.8', 'ISIC_6_hsv_f=0.9', 'ISIC_6_hsv_f=1.0',
                'ISIC_6_blur_f=0.1', 'ISIC_6_blur_f=0.2', 'ISIC_6_blur_f=0.3', 'ISIC_6_blur_f=0.4',
                'ISIC_6_blur_f=0.5', 'ISIC_6_blur_f=0.6', 'ISIC_6_blur_f=0.7',
                'ISIC_6_blur_f=0.8', 'ISIC_6_blur_f=0.9', 'ISIC_6_blur_f=1.0',
                'ISIC_6_small_random_f=0.1', 'ISIC_6_small_random_f=0.2', 'ISIC_6_small_random_f=0.3', 'ISIC_6_small_random_f=0.4',
                'ISIC_6_small_random_f=0.5', 'ISIC_6_small_random_f=0.6', 'ISIC_6_small_random_f=0.7',
                'ISIC_6_small_random_f=0.8', 'ISIC_6_small_random_f=0.9', 'ISIC_6_small_random_f=1.0',
                'ISIC_6_small_easy_f=0.1', 'ISIC_6_small_easy_f=0.2', 'ISIC_6_small_easy_f=0.3', 'ISIC_6_small_easy_f=0.4',
                'ISIC_6_small_easy_f=0.5', 'ISIC_6_small_easy_f=0.6', 'ISIC_6_small_easy_f=0.7',
                'ISIC_6_small_easy_f=0.8', 'ISIC_6_small_easy_f=0.9', 'ISIC_6_small_easy_f=1.0',
                'ISIC_6_small_hard_f=0.1', 'ISIC_6_small_hard_f=0.2', 'ISIC_6_small_hard_f=0.3', 'ISIC_6_small_hard_f=0.4',
                'ISIC_6_small_hard_f=0.5', 'ISIC_6_small_hard_f=0.6', 'ISIC_6_small_hard_f=0.7',
                'ISIC_6_small_hard_f=0.8', 'ISIC_6_small_hard_f=0.9', 'ISIC_6_small_hard_f=1.0',
                'ISIC_6_small_clusters_f=0.1', 'ISIC_6_small_clusters_f=0.2', 'ISIC_6_small_clusters_f=0.3', 'ISIC_6_small_clusters_f=0.4',
                'ISIC_6_small_clusters_f=0.5', 'ISIC_6_small_clusters_f=0.6', 'ISIC_6_small_clusters_f=0.7',
                'ISIC_6_small_clusters_f=0.8', 'ISIC_6_small_clusters_f=0.9', 'ISIC_6_small_clusters_f=1.0',
                'CNMC_2', 'CNMC_2_image_rot_f=0.1', 'CNMC_2_image_rot_f=0.2',
                'CNMC_2_image_rot_f=0.3', 'CNMC_2_image_rot_f=0.4', 'CNMC_2_image_rot_f=0.5',
                'CNMC_2_image_rot_f=0.6', 'CNMC_2_image_rot_f=0.7', 'CNMC_2_image_rot_f=0.8',
                'CNMC_2_image_rot_f=0.9', 'CNMC_2_image_rot_f=1.0', 'CNMC_2_image_translation_f=0.1',
                'CNMC_2_image_translation_f=0.2', 'CNMC_2_image_translation_f=0.3', 'CNMC_2_image_translation_f=0.4',
                'CNMC_2_image_translation_f=0.5', 'CNMC_2_image_translation_f=0.6', 'CNMC_2_image_translation_f=0.7',
                'CNMC_2_image_translation_f=0.8', 'CNMC_2_image_translation_f=0.9', 'CNMC_2_image_translation_f=1.0',
                'CNMC_2_image_zoom_f=0.1', 'CNMC_2_image_zoom_f=0.2', 'CNMC_2_image_zoom_f=0.3',
                'CNMC_2_image_zoom_f=0.4', 'CNMC_2_image_zoom_f=0.5', 'CNMC_2_image_zoom_f=0.6',
                'CNMC_2_image_zoom_f=0.7', 'CNMC_2_image_zoom_f=0.8', 'CNMC_2_image_zoom_f=0.9',
                'CNMC_2_image_zoom_f=1.0', 'CNMC_2_add_noise_gaussian_f=0.1', 'CNMC_2_add_noise_gaussian_f=0.2',
                'CNMC_2_add_noise_gaussian_f=0.3', 'CNMC_2_add_noise_gaussian_f=0.4', 'CNMC_2_add_noise_gaussian_f=0.5',
                'CNMC_2_add_noise_gaussian_f=0.6', 'CNMC_2_add_noise_gaussian_f=0.7', 'CNMC_2_add_noise_gaussian_f=0.8',
                'CNMC_2_add_noise_gaussian_f=0.9', 'CNMC_2_add_noise_gaussian_f=1.0', 'CNMC_2_add_noise_poisson_f=0.1',
                'CNMC_2_add_noise_poisson_f=0.2', 'CNMC_2_add_noise_poisson_f=0.3', 'CNMC_2_add_noise_poisson_f=0.4',
                'CNMC_2_add_noise_poisson_f=0.5', 'CNMC_2_add_noise_poisson_f=0.6', 'CNMC_2_add_noise_poisson_f=0.7',
                'CNMC_2_add_noise_poisson_f=0.8', 'CNMC_2_add_noise_poisson_f=0.9', 'CNMC_2_add_noise_poisson_f=1.0',
                'CNMC_2_add_noise_salt_and_pepper_f=0.1', 'CNMC_2_add_noise_salt_and_pepper_f=0.2',
                'CNMC_2_add_noise_salt_and_pepper_f=0.3', 'CNMC_2_add_noise_salt_and_pepper_f=0.4',
                'CNMC_2_add_noise_salt_and_pepper_f=0.5', 'CNMC_2_add_noise_salt_and_pepper_f=0.6',
                'CNMC_2_add_noise_salt_and_pepper_f=0.7', 'CNMC_2_add_noise_salt_and_pepper_f=0.8',
                'CNMC_2_add_noise_salt_and_pepper_f=0.9', 'CNMC_2_add_noise_salt_and_pepper_f=1.0',
                'CNMC_2_add_noise_speckle_f=0.1', 'CNMC_2_add_noise_speckle_f=0.2', 'CNMC_2_add_noise_speckle_f=0.3',
                'CNMC_2_add_noise_speckle_f=0.4', 'CNMC_2_add_noise_speckle_f=0.5', 'CNMC_2_add_noise_speckle_f=0.6',
                'CNMC_2_add_noise_speckle_f=0.7', 'CNMC_2_add_noise_speckle_f=0.8', 'CNMC_2_add_noise_speckle_f=0.9',
                'CNMC_2_add_noise_speckle_f=1.0', 'CNMC_2_imbalance_classes_f=0.1', 'CNMC_2_imbalance_classes_f=0.2',
                'CNMC_2_imbalance_classes_f=0.3', 'CNMC_2_imbalance_classes_f=0.4', 'CNMC_2_imbalance_classes_f=0.5',
                'CNMC_2_imbalance_classes_f=0.6', 'CNMC_2_imbalance_classes_f=0.7', 'CNMC_2_imbalance_classes_f=0.8',
                'CNMC_2_imbalance_classes_f=0.9', 'CNMC_2_imbalance_classes_f=1.0', 'CNMC_2_grayscale_f=0.1',
                'CNMC_2_grayscale_f=0.2', 'CNMC_2_grayscale_f=0.3', 'CNMC_2_grayscale_f=0.4',
                'CNMC_2_grayscale_f=0.5', 'CNMC_2_grayscale_f=0.6', 'CNMC_2_grayscale_f=0.7',
                'CNMC_2_grayscale_f=0.8', 'CNMC_2_grayscale_f=0.9', 'CNMC_2_grayscale_f=1.0',
                'CNMC_2_hsv_f=0.1', 'CNMC_2_hsv_f=0.2', 'CNMC_2_hsv_f=0.3', 'CNMC_2_hsv_f=0.4',
                'CNMC_2_hsv_f=0.5', 'CNMC_2_hsv_f=0.6', 'CNMC_2_hsv_f=0.7',
                'CNMC_2_hsv_f=0.8', 'CNMC_2_hsv_f=0.9', 'CNMC_2_hsv_f=1.0',
                'CNMC_2_blur_f=0.1', 'CNMC_2_blur_f=0.2', 'CNMC_2_blur_f=0.3', 'CNMC_2_blur_f=0.4',
                'CNMC_2_blur_f=0.5', 'CNMC_2_blur_f=0.6', 'CNMC_2_blur_f=0.7',
                'CNMC_2_blur_f=0.8', 'CNMC_2_blur_f=0.9', 'CNMC_2_blur_f=1.0',
                'CNMC_2_small_random_f=0.1', 'CNMC_2_small_random_f=0.2', 'CNMC_2_small_random_f=0.3', 'CNMC_2_small_random_f=0.4',
                'CNMC_2_small_random_f=0.5', 'CNMC_2_small_random_f=0.6', 'CNMC_2_small_random_f=0.7',
                'CNMC_2_small_random_f=0.8', 'CNMC_2_small_random_f=0.9', 'CNMC_2_small_random_f=1.0',
                'CNMC_2_small_easy_f=0.1', 'CNMC_2_small_easy_f=0.2', 'CNMC_2_small_easy_f=0.3', 'CNMC_2_small_easy_f=0.4',
                'CNMC_2_small_easy_f=0.5', 'CNMC_2_small_easy_f=0.6', 'CNMC_2_small_easy_f=0.7',
                'CNMC_2_small_easy_f=0.8', 'CNMC_2_small_easy_f=0.9', 'CNMC_2_small_easy_f=1.0',
                'CNMC_2_small_hard_f=0.1', 'CNMC_2_small_hard_f=0.2', 'CNMC_2_small_hard_f=0.3', 'CNMC_2_small_hard_f=0.4',
                'CNMC_2_small_hard_f=0.5', 'CNMC_2_small_hard_f=0.6', 'CNMC_2_small_hard_f=0.7',
                'CNMC_2_small_hard_f=0.8', 'CNMC_2_small_hard_f=0.9', 'CNMC_2_small_hard_f=1.0',
                'CNMC_2_small_clusters_f=0.1', 'CNMC_2_small_clusters_f=0.2', 'CNMC_2_small_clusters_f=0.3', 'CNMC_2_small_clusters_f=0.4',
                'CNMC_2_small_clusters_f=0.5', 'CNMC_2_small_clusters_f=0.6', 'CNMC_2_small_clusters_f=0.7',
                'CNMC_2_small_clusters_f=0.8', 'CNMC_2_small_clusters_f=0.9', 'CNMC_2_small_clusters_f=1.0',
                'CNMC_3', 'CNMC_3_image_rot_f=0.1', 'CNMC_3_image_rot_f=0.2',
                'CNMC_3_image_rot_f=0.3', 'CNMC_3_image_rot_f=0.4', 'CNMC_3_image_rot_f=0.5',
                'CNMC_3_image_rot_f=0.6', 'CNMC_3_image_rot_f=0.7', 'CNMC_3_image_rot_f=0.8',
                'CNMC_3_image_rot_f=0.9', 'CNMC_3_image_rot_f=1.0', 'CNMC_3_image_translation_f=0.1',
                'CNMC_3_image_translation_f=0.2', 'CNMC_3_image_translation_f=0.3', 'CNMC_3_image_translation_f=0.4',
                'CNMC_3_image_translation_f=0.5', 'CNMC_3_image_translation_f=0.6', 'CNMC_3_image_translation_f=0.7',
                'CNMC_3_image_translation_f=0.8', 'CNMC_3_image_translation_f=0.9', 'CNMC_3_image_translation_f=1.0',
                'CNMC_3_image_zoom_f=0.1', 'CNMC_3_image_zoom_f=0.2', 'CNMC_3_image_zoom_f=0.3',
                'CNMC_3_image_zoom_f=0.4', 'CNMC_3_image_zoom_f=0.5', 'CNMC_3_image_zoom_f=0.6',
                'CNMC_3_image_zoom_f=0.7', 'CNMC_3_image_zoom_f=0.8', 'CNMC_3_image_zoom_f=0.9',
                'CNMC_3_image_zoom_f=1.0', 'CNMC_3_add_noise_gaussian_f=0.1', 'CNMC_3_add_noise_gaussian_f=0.2',
                'CNMC_3_add_noise_gaussian_f=0.3', 'CNMC_3_add_noise_gaussian_f=0.4', 'CNMC_3_add_noise_gaussian_f=0.5',
                'CNMC_3_add_noise_gaussian_f=0.6', 'CNMC_3_add_noise_gaussian_f=0.7', 'CNMC_3_add_noise_gaussian_f=0.8',
                'CNMC_3_add_noise_gaussian_f=0.9', 'CNMC_3_add_noise_gaussian_f=1.0', 'CNMC_3_add_noise_poisson_f=0.1',
                'CNMC_3_add_noise_poisson_f=0.2', 'CNMC_3_add_noise_poisson_f=0.3', 'CNMC_3_add_noise_poisson_f=0.4',
                'CNMC_3_add_noise_poisson_f=0.5', 'CNMC_3_add_noise_poisson_f=0.6', 'CNMC_3_add_noise_poisson_f=0.7',
                'CNMC_3_add_noise_poisson_f=0.8', 'CNMC_3_add_noise_poisson_f=0.9', 'CNMC_3_add_noise_poisson_f=1.0',
                'CNMC_3_add_noise_salt_and_pepper_f=0.1', 'CNMC_3_add_noise_salt_and_pepper_f=0.2',
                'CNMC_3_add_noise_salt_and_pepper_f=0.3', 'CNMC_3_add_noise_salt_and_pepper_f=0.4',
                'CNMC_3_add_noise_salt_and_pepper_f=0.5', 'CNMC_3_add_noise_salt_and_pepper_f=0.6',
                'CNMC_3_add_noise_salt_and_pepper_f=0.7', 'CNMC_3_add_noise_salt_and_pepper_f=0.8',
                'CNMC_3_add_noise_salt_and_pepper_f=0.9', 'CNMC_3_add_noise_salt_and_pepper_f=1.0',
                'CNMC_3_add_noise_speckle_f=0.1', 'CNMC_3_add_noise_speckle_f=0.2', 'CNMC_3_add_noise_speckle_f=0.3',
                'CNMC_3_add_noise_speckle_f=0.4', 'CNMC_3_add_noise_speckle_f=0.5', 'CNMC_3_add_noise_speckle_f=0.6',
                'CNMC_3_add_noise_speckle_f=0.7', 'CNMC_3_add_noise_speckle_f=0.8', 'CNMC_3_add_noise_speckle_f=0.9',
                'CNMC_3_add_noise_speckle_f=1.0', 'CNMC_3_imbalance_classes_f=0.1', 'CNMC_3_imbalance_classes_f=0.2',
                'CNMC_3_imbalance_classes_f=0.3', 'CNMC_3_imbalance_classes_f=0.4', 'CNMC_3_imbalance_classes_f=0.5',
                'CNMC_3_imbalance_classes_f=0.6', 'CNMC_3_imbalance_classes_f=0.7', 'CNMC_3_imbalance_classes_f=0.8',
                'CNMC_3_imbalance_classes_f=0.9', 'CNMC_3_imbalance_classes_f=1.0', 'CNMC_3_grayscale_f=0.1',
                'CNMC_3_grayscale_f=0.2', 'CNMC_3_grayscale_f=0.3', 'CNMC_3_grayscale_f=0.4',
                'CNMC_3_grayscale_f=0.5', 'CNMC_3_grayscale_f=0.6', 'CNMC_3_grayscale_f=0.7',
                'CNMC_3_grayscale_f=0.8', 'CNMC_3_grayscale_f=0.9', 'CNMC_3_grayscale_f=1.0',
                'CNMC_3_hsv_f=0.1', 'CNMC_3_hsv_f=0.2', 'CNMC_3_hsv_f=0.3', 'CNMC_3_hsv_f=0.4',
                'CNMC_3_hsv_f=0.5', 'CNMC_3_hsv_f=0.6', 'CNMC_3_hsv_f=0.7',
                'CNMC_3_hsv_f=0.8', 'CNMC_3_hsv_f=0.9', 'CNMC_3_hsv_f=1.0',
                'CNMC_3_blur_f=0.1', 'CNMC_3_blur_f=0.2', 'CNMC_3_blur_f=0.3', 'CNMC_3_blur_f=0.4',
                'CNMC_3_blur_f=0.5', 'CNMC_3_blur_f=0.6', 'CNMC_3_blur_f=0.7',
                'CNMC_3_blur_f=0.8', 'CNMC_3_blur_f=0.9', 'CNMC_3_blur_f=1.0',
                'CNMC_3_small_random_f=0.1', 'CNMC_3_small_random_f=0.2', 'CNMC_3_small_random_f=0.3', 'CNMC_3_small_random_f=0.4',
                'CNMC_3_small_random_f=0.5', 'CNMC_3_small_random_f=0.6', 'CNMC_3_small_random_f=0.7',
                'CNMC_3_small_random_f=0.8', 'CNMC_3_small_random_f=0.9', 'CNMC_3_small_random_f=1.0',
                'CNMC_3_small_easy_f=0.1', 'CNMC_3_small_easy_f=0.2', 'CNMC_3_small_easy_f=0.3', 'CNMC_3_small_easy_f=0.4',
                'CNMC_3_small_easy_f=0.5', 'CNMC_3_small_easy_f=0.6', 'CNMC_3_small_easy_f=0.7',
                'CNMC_3_small_easy_f=0.8', 'CNMC_3_small_easy_f=0.9', 'CNMC_3_small_easy_f=1.0',
                'CNMC_3_small_hard_f=0.1', 'CNMC_3_small_hard_f=0.2', 'CNMC_3_small_hard_f=0.3', 'CNMC_3_small_hard_f=0.4',
                'CNMC_3_small_hard_f=0.5', 'CNMC_3_small_hard_f=0.6', 'CNMC_3_small_hard_f=0.7',
                'CNMC_3_small_hard_f=0.8', 'CNMC_3_small_hard_f=0.9', 'CNMC_3_small_hard_f=1.0',
                'CNMC_3_small_clusters_f=0.1', 'CNMC_3_small_clusters_f=0.2', 'CNMC_3_small_clusters_f=0.3', 'CNMC_3_small_clusters_f=0.4',
                'CNMC_3_small_clusters_f=0.5', 'CNMC_3_small_clusters_f=0.6', 'CNMC_3_small_clusters_f=0.7',
                'CNMC_3_small_clusters_f=0.8', 'CNMC_3_small_clusters_f=0.9', 'CNMC_3_small_clusters_f=1.0',
                'CNMC_4', 'CNMC_4_image_rot_f=0.1', 'CNMC_4_image_rot_f=0.2',
                'CNMC_4_image_rot_f=0.3', 'CNMC_4_image_rot_f=0.4', 'CNMC_4_image_rot_f=0.5',
                'CNMC_4_image_rot_f=0.6', 'CNMC_4_image_rot_f=0.7', 'CNMC_4_image_rot_f=0.8',
                'CNMC_4_image_rot_f=0.9', 'CNMC_4_image_rot_f=1.0', 'CNMC_4_image_translation_f=0.1',
                'CNMC_4_image_translation_f=0.2', 'CNMC_4_image_translation_f=0.3', 'CNMC_4_image_translation_f=0.4',
                'CNMC_4_image_translation_f=0.5', 'CNMC_4_image_translation_f=0.6', 'CNMC_4_image_translation_f=0.7',
                'CNMC_4_image_translation_f=0.8', 'CNMC_4_image_translation_f=0.9', 'CNMC_4_image_translation_f=1.0',
                'CNMC_4_image_zoom_f=0.1', 'CNMC_4_image_zoom_f=0.2', 'CNMC_4_image_zoom_f=0.3',
                'CNMC_4_image_zoom_f=0.4', 'CNMC_4_image_zoom_f=0.5', 'CNMC_4_image_zoom_f=0.6',
                'CNMC_4_image_zoom_f=0.7', 'CNMC_4_image_zoom_f=0.8', 'CNMC_4_image_zoom_f=0.9',
                'CNMC_4_image_zoom_f=1.0', 'CNMC_4_add_noise_gaussian_f=0.1', 'CNMC_4_add_noise_gaussian_f=0.2',
                'CNMC_4_add_noise_gaussian_f=0.3', 'CNMC_4_add_noise_gaussian_f=0.4', 'CNMC_4_add_noise_gaussian_f=0.5',
                'CNMC_4_add_noise_gaussian_f=0.6', 'CNMC_4_add_noise_gaussian_f=0.7', 'CNMC_4_add_noise_gaussian_f=0.8',
                'CNMC_4_add_noise_gaussian_f=0.9', 'CNMC_4_add_noise_gaussian_f=1.0', 'CNMC_4_add_noise_poisson_f=0.1',
                'CNMC_4_add_noise_poisson_f=0.2', 'CNMC_4_add_noise_poisson_f=0.3', 'CNMC_4_add_noise_poisson_f=0.4',
                'CNMC_4_add_noise_poisson_f=0.5', 'CNMC_4_add_noise_poisson_f=0.6', 'CNMC_4_add_noise_poisson_f=0.7',
                'CNMC_4_add_noise_poisson_f=0.8', 'CNMC_4_add_noise_poisson_f=0.9', 'CNMC_4_add_noise_poisson_f=1.0',
                'CNMC_4_add_noise_salt_and_pepper_f=0.1', 'CNMC_4_add_noise_salt_and_pepper_f=0.2',
                'CNMC_4_add_noise_salt_and_pepper_f=0.3', 'CNMC_4_add_noise_salt_and_pepper_f=0.4',
                'CNMC_4_add_noise_salt_and_pepper_f=0.5', 'CNMC_4_add_noise_salt_and_pepper_f=0.6',
                'CNMC_4_add_noise_salt_and_pepper_f=0.7', 'CNMC_4_add_noise_salt_and_pepper_f=0.8',
                'CNMC_4_add_noise_salt_and_pepper_f=0.9', 'CNMC_4_add_noise_salt_and_pepper_f=1.0',
                'CNMC_4_add_noise_speckle_f=0.1', 'CNMC_4_add_noise_speckle_f=0.2', 'CNMC_4_add_noise_speckle_f=0.3',
                'CNMC_4_add_noise_speckle_f=0.4', 'CNMC_4_add_noise_speckle_f=0.5', 'CNMC_4_add_noise_speckle_f=0.6',
                'CNMC_4_add_noise_speckle_f=0.7', 'CNMC_4_add_noise_speckle_f=0.8', 'CNMC_4_add_noise_speckle_f=0.9',
                'CNMC_4_add_noise_speckle_f=1.0', 'CNMC_4_imbalance_classes_f=0.1', 'CNMC_4_imbalance_classes_f=0.2',
                'CNMC_4_imbalance_classes_f=0.3', 'CNMC_4_imbalance_classes_f=0.4', 'CNMC_4_imbalance_classes_f=0.5',
                'CNMC_4_imbalance_classes_f=0.6', 'CNMC_4_imbalance_classes_f=0.7', 'CNMC_4_imbalance_classes_f=0.8',
                'CNMC_4_imbalance_classes_f=0.9', 'CNMC_4_imbalance_classes_f=1.0', 'CNMC_4_grayscale_f=0.1',
                'CNMC_4_grayscale_f=0.2', 'CNMC_4_grayscale_f=0.3', 'CNMC_4_grayscale_f=0.4',
                'CNMC_4_grayscale_f=0.5', 'CNMC_4_grayscale_f=0.6', 'CNMC_4_grayscale_f=0.7',
                'CNMC_4_grayscale_f=0.8', 'CNMC_4_grayscale_f=0.9', 'CNMC_4_grayscale_f=1.0',
                'CNMC_4_hsv_f=0.1', 'CNMC_4_hsv_f=0.2', 'CNMC_4_hsv_f=0.3', 'CNMC_4_hsv_f=0.4',
                'CNMC_4_hsv_f=0.5', 'CNMC_4_hsv_f=0.6', 'CNMC_4_hsv_f=0.7',
                'CNMC_4_hsv_f=0.8', 'CNMC_4_hsv_f=0.9', 'CNMC_4_hsv_f=1.0',
                'CNMC_4_blur_f=0.1', 'CNMC_4_blur_f=0.2', 'CNMC_4_blur_f=0.3', 'CNMC_4_blur_f=0.4',
                'CNMC_4_blur_f=0.5', 'CNMC_4_blur_f=0.6', 'CNMC_4_blur_f=0.7',
                'CNMC_4_blur_f=0.8', 'CNMC_4_blur_f=0.9', 'CNMC_4_blur_f=1.0',
                'CNMC_4_small_random_f=0.1', 'CNMC_4_small_random_f=0.2', 'CNMC_4_small_random_f=0.3', 'CNMC_4_small_random_f=0.4',
                'CNMC_4_small_random_f=0.5', 'CNMC_4_small_random_f=0.6', 'CNMC_4_small_random_f=0.7',
                'CNMC_4_small_random_f=0.8', 'CNMC_4_small_random_f=0.9', 'CNMC_4_small_random_f=1.0',
                'CNMC_4_small_easy_f=0.1', 'CNMC_4_small_easy_f=0.2', 'CNMC_4_small_easy_f=0.3', 'CNMC_4_small_easy_f=0.4',
                'CNMC_4_small_easy_f=0.5', 'CNMC_4_small_easy_f=0.6', 'CNMC_4_small_easy_f=0.7',
                'CNMC_4_small_easy_f=0.8', 'CNMC_4_small_easy_f=0.9', 'CNMC_4_small_easy_f=1.0',
                'CNMC_4_small_hard_f=0.1', 'CNMC_4_small_hard_f=0.2', 'CNMC_4_small_hard_f=0.3', 'CNMC_4_small_hard_f=0.4',
                'CNMC_4_small_hard_f=0.5', 'CNMC_4_small_hard_f=0.6', 'CNMC_4_small_hard_f=0.7',
                'CNMC_4_small_hard_f=0.8', 'CNMC_4_small_hard_f=0.9', 'CNMC_4_small_hard_f=1.0',
                'CNMC_4_small_clusters_f=0.1', 'CNMC_4_small_clusters_f=0.2', 'CNMC_4_small_clusters_f=0.3', 'CNMC_4_small_clusters_f=0.4',
                'CNMC_4_small_clusters_f=0.5', 'CNMC_4_small_clusters_f=0.6', 'CNMC_4_small_clusters_f=0.7',
                'CNMC_4_small_clusters_f=0.8', 'CNMC_4_small_clusters_f=0.9', 'CNMC_4_small_clusters_f=1.0',
                'CNMC_5', 'CNMC_5_image_rot_f=0.1', 'CNMC_5_image_rot_f=0.2',
                'CNMC_5_image_rot_f=0.3', 'CNMC_5_image_rot_f=0.4', 'CNMC_5_image_rot_f=0.5',
                'CNMC_5_image_rot_f=0.6', 'CNMC_5_image_rot_f=0.7', 'CNMC_5_image_rot_f=0.8',
                'CNMC_5_image_rot_f=0.9', 'CNMC_5_image_rot_f=1.0', 'CNMC_5_image_translation_f=0.1',
                'CNMC_5_image_translation_f=0.2', 'CNMC_5_image_translation_f=0.3', 'CNMC_5_image_translation_f=0.4',
                'CNMC_5_image_translation_f=0.5', 'CNMC_5_image_translation_f=0.6', 'CNMC_5_image_translation_f=0.7',
                'CNMC_5_image_translation_f=0.8', 'CNMC_5_image_translation_f=0.9', 'CNMC_5_image_translation_f=1.0',
                'CNMC_5_image_zoom_f=0.1', 'CNMC_5_image_zoom_f=0.2', 'CNMC_5_image_zoom_f=0.3',
                'CNMC_5_image_zoom_f=0.4', 'CNMC_5_image_zoom_f=0.5', 'CNMC_5_image_zoom_f=0.6',
                'CNMC_5_image_zoom_f=0.7', 'CNMC_5_image_zoom_f=0.8', 'CNMC_5_image_zoom_f=0.9',
                'CNMC_5_image_zoom_f=1.0', 'CNMC_5_add_noise_gaussian_f=0.1', 'CNMC_5_add_noise_gaussian_f=0.2',
                'CNMC_5_add_noise_gaussian_f=0.3', 'CNMC_5_add_noise_gaussian_f=0.4', 'CNMC_5_add_noise_gaussian_f=0.5',
                'CNMC_5_add_noise_gaussian_f=0.6', 'CNMC_5_add_noise_gaussian_f=0.7', 'CNMC_5_add_noise_gaussian_f=0.8',
                'CNMC_5_add_noise_gaussian_f=0.9', 'CNMC_5_add_noise_gaussian_f=1.0', 'CNMC_5_add_noise_poisson_f=0.1',
                'CNMC_5_add_noise_poisson_f=0.2', 'CNMC_5_add_noise_poisson_f=0.3', 'CNMC_5_add_noise_poisson_f=0.4',
                'CNMC_5_add_noise_poisson_f=0.5', 'CNMC_5_add_noise_poisson_f=0.6', 'CNMC_5_add_noise_poisson_f=0.7',
                'CNMC_5_add_noise_poisson_f=0.8', 'CNMC_5_add_noise_poisson_f=0.9', 'CNMC_5_add_noise_poisson_f=1.0',
                'CNMC_5_add_noise_salt_and_pepper_f=0.1', 'CNMC_5_add_noise_salt_and_pepper_f=0.2',
                'CNMC_5_add_noise_salt_and_pepper_f=0.3', 'CNMC_5_add_noise_salt_and_pepper_f=0.4',
                'CNMC_5_add_noise_salt_and_pepper_f=0.5', 'CNMC_5_add_noise_salt_and_pepper_f=0.6',
                'CNMC_5_add_noise_salt_and_pepper_f=0.7', 'CNMC_5_add_noise_salt_and_pepper_f=0.8',
                'CNMC_5_add_noise_salt_and_pepper_f=0.9', 'CNMC_5_add_noise_salt_and_pepper_f=1.0',
                'CNMC_5_add_noise_speckle_f=0.1', 'CNMC_5_add_noise_speckle_f=0.2', 'CNMC_5_add_noise_speckle_f=0.3',
                'CNMC_5_add_noise_speckle_f=0.4', 'CNMC_5_add_noise_speckle_f=0.5', 'CNMC_5_add_noise_speckle_f=0.6',
                'CNMC_5_add_noise_speckle_f=0.7', 'CNMC_5_add_noise_speckle_f=0.8', 'CNMC_5_add_noise_speckle_f=0.9',
                'CNMC_5_add_noise_speckle_f=1.0', 'CNMC_5_imbalance_classes_f=0.1', 'CNMC_5_imbalance_classes_f=0.2',
                'CNMC_5_imbalance_classes_f=0.3', 'CNMC_5_imbalance_classes_f=0.4', 'CNMC_5_imbalance_classes_f=0.5',
                'CNMC_5_imbalance_classes_f=0.6', 'CNMC_5_imbalance_classes_f=0.7', 'CNMC_5_imbalance_classes_f=0.8',
                'CNMC_5_imbalance_classes_f=0.9', 'CNMC_5_imbalance_classes_f=1.0', 'CNMC_5_grayscale_f=0.1',
                'CNMC_5_grayscale_f=0.2', 'CNMC_5_grayscale_f=0.3', 'CNMC_5_grayscale_f=0.4',
                'CNMC_5_grayscale_f=0.5', 'CNMC_5_grayscale_f=0.6', 'CNMC_5_grayscale_f=0.7',
                'CNMC_5_grayscale_f=0.8', 'CNMC_5_grayscale_f=0.9', 'CNMC_5_grayscale_f=1.0',
                'CNMC_5_hsv_f=0.1', 'CNMC_5_hsv_f=0.2', 'CNMC_5_hsv_f=0.3', 'CNMC_5_hsv_f=0.4',
                'CNMC_5_hsv_f=0.5', 'CNMC_5_hsv_f=0.6', 'CNMC_5_hsv_f=0.7',
                'CNMC_5_hsv_f=0.8', 'CNMC_5_hsv_f=0.9', 'CNMC_5_hsv_f=1.0',
                'CNMC_5_blur_f=0.1', 'CNMC_5_blur_f=0.2', 'CNMC_5_blur_f=0.3', 'CNMC_5_blur_f=0.4',
                'CNMC_5_blur_f=0.5', 'CNMC_5_blur_f=0.6', 'CNMC_5_blur_f=0.7',
                'CNMC_5_blur_f=0.8', 'CNMC_5_blur_f=0.9', 'CNMC_5_blur_f=1.0',
                'CNMC_5_small_random_f=0.1', 'CNMC_5_small_random_f=0.2', 'CNMC_5_small_random_f=0.3', 'CNMC_5_small_random_f=0.4',
                'CNMC_5_small_random_f=0.5', 'CNMC_5_small_random_f=0.6', 'CNMC_5_small_random_f=0.7',
                'CNMC_5_small_random_f=0.8', 'CNMC_5_small_random_f=0.9', 'CNMC_5_small_random_f=1.0',
                'CNMC_5_small_easy_f=0.1', 'CNMC_5_small_easy_f=0.2', 'CNMC_5_small_easy_f=0.3', 'CNMC_5_small_easy_f=0.4',
                'CNMC_5_small_easy_f=0.5', 'CNMC_5_small_easy_f=0.6', 'CNMC_5_small_easy_f=0.7',
                'CNMC_5_small_easy_f=0.8', 'CNMC_5_small_easy_f=0.9', 'CNMC_5_small_easy_f=1.0',
                'CNMC_5_small_hard_f=0.1', 'CNMC_5_small_hard_f=0.2', 'CNMC_5_small_hard_f=0.3', 'CNMC_5_small_hard_f=0.4',
                'CNMC_5_small_hard_f=0.5', 'CNMC_5_small_hard_f=0.6', 'CNMC_5_small_hard_f=0.7',
                'CNMC_5_small_hard_f=0.8', 'CNMC_5_small_hard_f=0.9', 'CNMC_5_small_hard_f=1.0',
                'CNMC_5_small_clusters_f=0.1', 'CNMC_5_small_clusters_f=0.2', 'CNMC_5_small_clusters_f=0.3', 'CNMC_5_small_clusters_f=0.4',
                'CNMC_5_small_clusters_f=0.5', 'CNMC_5_small_clusters_f=0.6', 'CNMC_5_small_clusters_f=0.7',
                'CNMC_5_small_clusters_f=0.8', 'CNMC_5_small_clusters_f=0.9', 'CNMC_5_small_clusters_f=1.0',
                'CNMC_6', 'CNMC_6_image_rot_f=0.1', 'CNMC_6_image_rot_f=0.2',
                'CNMC_6_image_rot_f=0.3', 'CNMC_6_image_rot_f=0.4', 'CNMC_6_image_rot_f=0.5',
                'CNMC_6_image_rot_f=0.6', 'CNMC_6_image_rot_f=0.7', 'CNMC_6_image_rot_f=0.8',
                'CNMC_6_image_rot_f=0.9', 'CNMC_6_image_rot_f=1.0', 'CNMC_6_image_translation_f=0.1',
                'CNMC_6_image_translation_f=0.2', 'CNMC_6_image_translation_f=0.3', 'CNMC_6_image_translation_f=0.4',
                'CNMC_6_image_translation_f=0.5', 'CNMC_6_image_translation_f=0.6', 'CNMC_6_image_translation_f=0.7',
                'CNMC_6_image_translation_f=0.8', 'CNMC_6_image_translation_f=0.9', 'CNMC_6_image_translation_f=1.0',
                'CNMC_6_image_zoom_f=0.1', 'CNMC_6_image_zoom_f=0.2', 'CNMC_6_image_zoom_f=0.3',
                'CNMC_6_image_zoom_f=0.4', 'CNMC_6_image_zoom_f=0.5', 'CNMC_6_image_zoom_f=0.6',
                'CNMC_6_image_zoom_f=0.7', 'CNMC_6_image_zoom_f=0.8', 'CNMC_6_image_zoom_f=0.9',
                'CNMC_6_image_zoom_f=1.0', 'CNMC_6_add_noise_gaussian_f=0.1', 'CNMC_6_add_noise_gaussian_f=0.2',
                'CNMC_6_add_noise_gaussian_f=0.3', 'CNMC_6_add_noise_gaussian_f=0.4', 'CNMC_6_add_noise_gaussian_f=0.5',
                'CNMC_6_add_noise_gaussian_f=0.6', 'CNMC_6_add_noise_gaussian_f=0.7', 'CNMC_6_add_noise_gaussian_f=0.8',
                'CNMC_6_add_noise_gaussian_f=0.9', 'CNMC_6_add_noise_gaussian_f=1.0', 'CNMC_6_add_noise_poisson_f=0.1',
                'CNMC_6_add_noise_poisson_f=0.2', 'CNMC_6_add_noise_poisson_f=0.3', 'CNMC_6_add_noise_poisson_f=0.4',
                'CNMC_6_add_noise_poisson_f=0.5', 'CNMC_6_add_noise_poisson_f=0.6', 'CNMC_6_add_noise_poisson_f=0.7',
                'CNMC_6_add_noise_poisson_f=0.8', 'CNMC_6_add_noise_poisson_f=0.9', 'CNMC_6_add_noise_poisson_f=1.0',
                'CNMC_6_add_noise_salt_and_pepper_f=0.1', 'CNMC_6_add_noise_salt_and_pepper_f=0.2',
                'CNMC_6_add_noise_salt_and_pepper_f=0.3', 'CNMC_6_add_noise_salt_and_pepper_f=0.4',
                'CNMC_6_add_noise_salt_and_pepper_f=0.5', 'CNMC_6_add_noise_salt_and_pepper_f=0.6',
                'CNMC_6_add_noise_salt_and_pepper_f=0.7', 'CNMC_6_add_noise_salt_and_pepper_f=0.8',
                'CNMC_6_add_noise_salt_and_pepper_f=0.9', 'CNMC_6_add_noise_salt_and_pepper_f=1.0',
                'CNMC_6_add_noise_speckle_f=0.1', 'CNMC_6_add_noise_speckle_f=0.2', 'CNMC_6_add_noise_speckle_f=0.3',
                'CNMC_6_add_noise_speckle_f=0.4', 'CNMC_6_add_noise_speckle_f=0.5', 'CNMC_6_add_noise_speckle_f=0.6',
                'CNMC_6_add_noise_speckle_f=0.7', 'CNMC_6_add_noise_speckle_f=0.8', 'CNMC_6_add_noise_speckle_f=0.9',
                'CNMC_6_add_noise_speckle_f=1.0', 'CNMC_6_imbalance_classes_f=0.1', 'CNMC_6_imbalance_classes_f=0.2',
                'CNMC_6_imbalance_classes_f=0.3', 'CNMC_6_imbalance_classes_f=0.4', 'CNMC_6_imbalance_classes_f=0.5',
                'CNMC_6_imbalance_classes_f=0.6', 'CNMC_6_imbalance_classes_f=0.7', 'CNMC_6_imbalance_classes_f=0.8',
                'CNMC_6_imbalance_classes_f=0.9', 'CNMC_6_imbalance_classes_f=1.0', 'CNMC_6_grayscale_f=0.1',
                'CNMC_6_grayscale_f=0.2', 'CNMC_6_grayscale_f=0.3', 'CNMC_6_grayscale_f=0.4',
                'CNMC_6_grayscale_f=0.5', 'CNMC_6_grayscale_f=0.6', 'CNMC_6_grayscale_f=0.7',
                'CNMC_6_grayscale_f=0.8', 'CNMC_6_grayscale_f=0.9', 'CNMC_6_grayscale_f=1.0',
                'CNMC_6_hsv_f=0.1', 'CNMC_6_hsv_f=0.2', 'CNMC_6_hsv_f=0.3', 'CNMC_6_hsv_f=0.4',
                'CNMC_6_hsv_f=0.5', 'CNMC_6_hsv_f=0.6', 'CNMC_6_hsv_f=0.7',
                'CNMC_6_hsv_f=0.8', 'CNMC_6_hsv_f=0.9', 'CNMC_6_hsv_f=1.0',
                'CNMC_6_blur_f=0.1', 'CNMC_6_blur_f=0.2', 'CNMC_6_blur_f=0.3', 'CNMC_6_blur_f=0.4',
                'CNMC_6_blur_f=0.5', 'CNMC_6_blur_f=0.6', 'CNMC_6_blur_f=0.7',
                'CNMC_6_blur_f=0.8', 'CNMC_6_blur_f=0.9', 'CNMC_6_blur_f=1.0',
                'CNMC_6_small_random_f=0.1', 'CNMC_6_small_random_f=0.2', 'CNMC_6_small_random_f=0.3', 'CNMC_6_small_random_f=0.4',
                'CNMC_6_small_random_f=0.5', 'CNMC_6_small_random_f=0.6', 'CNMC_6_small_random_f=0.7',
                'CNMC_6_small_random_f=0.8', 'CNMC_6_small_random_f=0.9', 'CNMC_6_small_random_f=1.0',
                'CNMC_6_small_easy_f=0.1', 'CNMC_6_small_easy_f=0.2', 'CNMC_6_small_easy_f=0.3', 'CNMC_6_small_easy_f=0.4',
                'CNMC_6_small_easy_f=0.5', 'CNMC_6_small_easy_f=0.6', 'CNMC_6_small_easy_f=0.7',
                'CNMC_6_small_easy_f=0.8', 'CNMC_6_small_easy_f=0.9', 'CNMC_6_small_easy_f=1.0',
                'CNMC_6_small_hard_f=0.1', 'CNMC_6_small_hard_f=0.2', 'CNMC_6_small_hard_f=0.3', 'CNMC_6_small_hard_f=0.4',
                'CNMC_6_small_hard_f=0.5', 'CNMC_6_small_hard_f=0.6', 'CNMC_6_small_hard_f=0.7',
                'CNMC_6_small_hard_f=0.8', 'CNMC_6_small_hard_f=0.9', 'CNMC_6_small_hard_f=1.0',
                'CNMC_6_small_clusters_f=0.1', 'CNMC_6_small_clusters_f=0.2', 'CNMC_6_small_clusters_f=0.3', 'CNMC_6_small_clusters_f=0.4',
                'CNMC_6_small_clusters_f=0.5', 'CNMC_6_small_clusters_f=0.6', 'CNMC_6_small_clusters_f=0.7',
                'CNMC_6_small_clusters_f=0.8', 'CNMC_6_small_clusters_f=0.9', 'CNMC_6_small_clusters_f=1.0'],
        required='fine_tuning' in sys.argv or 'SVM' in sys.argv,
        help='source dataset, only when using fine_tuning or SVM, name should be in same format as target dataset',
        metavar='source dataset, only when using fine_tuning or SVM, name should be in same format as target dataset')
    parser.add_argument('-bs', '--batchsize', default=32, help='batch size')
    args = vars(parser.parse_args())

    main()
