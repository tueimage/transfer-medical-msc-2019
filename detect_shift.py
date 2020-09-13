import os
import argparse
import json
import sys
import glob
import numpy as np
import pandas as pd
import cv2
from models import get_shiftCNN
from keras.optimizers import SGD
from utils import *
from scipy.stats import binom_test, ks_2samp
from openpyxl import load_workbook, Workbook

# choose GPU
os.environ["CUDA_VISIBLE_DEVICES"] = "2"

# construct argument parser and parse the arguments
parser = argparse.ArgumentParser()
parser.add_argument('-m',
    '--mode',
    choices=['train_val', 'datasets'],
    required=True,
    help='dataset shift to detect, train_val for shift between train and validation set')
parser.add_argument('-d',
    '--dataset',
    required=True,
    help='dataset to use')
parser.add_argument('-d2',
    '--dataset2',
    required='datasets' in sys.argv,
    help='second dataset to use')

args = vars(parser.parse_args())

# read parameters for wanted dataset from config file
with open('config.json') as json_file:
    file = json.load(json_file)
    config = file[args['dataset']]
    config2 = file[args['dataset2']]

# create results directory if it doesn't exist
resultspath = os.path.join(os.path.dirname(os.getcwd()), 'results')
if not os.path.exists(resultspath):
    os.makedirs(resultspath)

if args['mode'] == 'train_val':
    # get training and validation path
    trainingpath = config['trainingpath']
    validationpath = config['validationpath']

    # get image paths
    impaths_train = glob.glob(os.path.join(trainingpath, '**/*.jpg'))
    impaths_val = glob.glob(os.path.join(validationpath, '**/*.jpg'))

    # now split both datasplits in half to use for training and testing
    impaths_train_t = impaths_train[:int(np.ceil(0.5*len(impaths_train)))]
    impaths_val_t = impaths_val[:int(np.ceil(0.5*len(impaths_val)))]

    impaths_train_s = impaths_train[int(np.ceil(0.5*len(impaths_train))):]
    impaths_val_s = impaths_val[int(np.ceil(0.5*len(impaths_val))):]

if args['mode'] == 'datasets':
    trainingpath1 = config['trainingpath']
    trainingpath2 = config2['trainingpath']

    impaths_1 = glob.glob(os.path.join(trainingpath1, '**/*.jpg'))
    impaths_2 = glob.glob(os.path.join(trainingpath2, '**/*.jpg'))

    # now split both datasplits in half to use for training and testing
    impaths_train_t = impaths_1[:int(np.ceil(0.5*len(impaths_1)))]
    impaths_val_t = impaths_2[:int(np.ceil(0.5*len(impaths_2)))]

    impaths_train_s = impaths_1[int(np.ceil(0.5*len(impaths_1))):]
    impaths_val_s = impaths_2[int(np.ceil(0.5*len(impaths_2))):]

# create arrays with training and validation data
x_train, y_train = [], []
for path in impaths_train_t:
    x_train.append(cv2.imread(path))
    y_train.append(0)
for path in impaths_val_t:
    x_train.append(cv2.imread(path))
    y_train.append(1)

x_train = np.array(x_train)
y_train = np.array(y_train)

x_val, y_val = [], []
for path in impaths_train_s:
    x_val.append(cv2.imread(path))
    y_val.append(0)
for path in impaths_val_s:
    x_val.append(cv2.imread(path))
    y_val.append(1)

x_train = np.asarray(x_train)
y_train = np.asarray(y_train)
x_val = np.asarray(x_val)
y_val = np.asarray(y_val)

# convert arrays to float32
x_train = x_train.astype('float32')
x_val = x_val.astype('float32')

# scale the data between 0 and 1
x_train /= 255.0
x_val /= 255.0

# get shift detection CNN
CNN = get_shiftCNN(input_shape=(224, 224, 3))
CNN.summary()

# compile CNN
sgd = SGD(lr=1e-5, nesterov=True)
CNN.compile(optimizer=sgd, loss='binary_crossentropy', metrics=['accuracy'])

# train shift detection classifier
hist = CNN.fit(x_train, y_train, batch_size=32, epochs=10, verbose=1, validation_data=(x_val, y_val))

# now make predictions
print("evaluating model...")
preds = CNN.predict(x_val, batch_size=1, verbose=1)

# preds is an array like [[x] [x] [x]], make it into array like [x x x]
preds = np.asarray([label for sublist in preds for label in sublist])

# get true labels
true_labels = y_val

# calculate AUC
fpr, tpr, thresholds, AUC = AUC_score(preds, true_labels)

# calculate accuracy score
acc = accuracy(preds, true_labels)

# create binary predictions
pred_labels = np.where(preds > 0.5, 1, 0).astype(int)

unique, counts = np.unique(true_labels, return_counts=True)
print("distribution of true labels: {}".format(dict(zip(unique, counts))))

unique, counts = np.unique(pred_labels, return_counts=True)
print("distribution of predicted labels: {}".format(dict(zip(unique, counts))))

print("accuracy: {}".format(acc))

absAUC = round(abs(AUC-0.5), 3)

# create a savepath for results and create a sheet to avoid errors
savefile = os.path.join(config['output_path'], 'results_shift_CNN_3.xlsx')

if not os.path.exists(config['output_path']):
    os.makedirs(config['output_path'])

# also create excel file already to avoid errors, if it doesn't exist yet
if not os.path.exists(savefile):
    Workbook().save(savefile)

# save in path
test_results = pd.Series({'dataset2': args['dataset2'], 'absAUC': absAUC, 'AUC': AUC, 'acc': acc})

# read existing rows and add test results
try:
    df = pd.read_excel(savefile, sheet_name='detect_shift')
except:
    df = pd.DataFrame(columns=['dataset2', 'absAUC', 'AUC', 'acc'])

df = df.append(test_results, ignore_index=True)
df.set_index('dataset2', inplace=True)

# save results in excel file
with pd.ExcelWriter(savefile, engine='openpyxl') as writer:
    writer.book = load_workbook(savefile)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    df.index.name = 'dataset2'
    df.to_excel(writer, sheet_name='detect_shift')
